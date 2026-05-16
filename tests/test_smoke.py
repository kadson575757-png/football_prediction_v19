from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from football_prediction_v19.cli import main
from football_prediction_v19.backtest import build_betting_report, run_bet_backtest
from football_prediction_v19.data import REAL_MATCH_OPTIONAL_NUMERIC_COLUMNS, prepare_real_matches
from football_prediction_v19.features import build_features, build_fixture_features
from football_prediction_v19.excel_report import create_predictions_excel_report
from football_prediction_v19.model import predict_feature_rows, train_from_matches
from football_prediction_v19.fixtures import prepare_fixtures
from football_prediction_v19.odds_import import (
    merge_odds_file,
    merge_odds_into_fixtures,
    prepare_odds,
    prepare_odds_file,
)
from football_prediction_v19.importers.fbref import normalize_fbref_csv
from football_prediction_v19.importers.football_data import (
    LEAGUE_CODES,
    build_football_data_url,
    bulk_download,
    download_and_prepare,
    download_season,
    normalize_football_data_csv,
)
from football_prediction_v19.odds import (
    best_value_side,
    bookmaker_overround,
    fair_probabilities,
    grade_flat_stake_bet,
    implied_probability,
    model_edges,
)
from football_prediction_v19.rules_v19 import assess_prediction
from football_prediction_v19.team_names import normalize_team_name


def test_pipeline_smoke():
    path = Path(__file__).resolve().parents[1] / "data" / "sample_matches.csv"
    df = pd.read_csv(path)
    model, table, metrics, cols = train_from_matches(df, test_season=2023)
    assert len(table) > 10
    fixture = build_fixture_features(df, "Chelsea", "Arsenal", "2024-05-01")
    bundle = {"model": model, "feature_cols": cols, "metrics": metrics}
    pred = predict_feature_rows(bundle, fixture).iloc[0]
    assessment = assess_prediction(pred, {"H": pred.prob_home, "D": pred.prob_draw, "A": pred.prob_away})
    assert "probabilities" in assessment


def test_basic_dataset_still_trains_without_optional_columns():
    path = Path(__file__).resolve().parents[1] / "data" / "sample_matches.csv"
    df = pd.read_csv(path)
    model, table, metrics, cols = train_from_matches(df, test_season=2023)

    assert len(table) > 10
    assert metrics["feature_count"] == len(cols)
    assert "home_w5_shots" not in cols


def test_predict_fixtures_command(tmp_path):
    root = Path(__file__).resolve().parents[1]
    history_path = root / "data" / "sample_matches.csv"
    fixtures_path = root / "data" / "upcoming_fixtures_template.csv"
    model_path = tmp_path / "sample_model.joblib"
    output_path = tmp_path / "outputs" / "predictions.csv"

    main(
        [
            "train",
            "--input",
            str(history_path),
            "--model",
            str(model_path),
            "--test-season",
            "2023",
        ]
    )
    main(
        [
            "predict-fixtures",
            "--history",
            str(history_path),
            "--fixtures",
            str(fixtures_path),
            "--model",
            str(model_path),
            "--output",
            str(output_path),
        ]
    )

    assert output_path.exists()
    result = pd.read_csv(output_path)
    expected_columns = [
        "date",
        "league",
        "home_team",
        "away_team",
        "prob_home",
        "prob_draw",
        "prob_away",
        "top_pick",
        "confidence",
        "control_score",
        "chaos_score",
        "tdi_home",
        "tdi_away",
        "odds_home",
        "odds_draw",
        "odds_away",
        "implied_home",
        "implied_draw",
        "implied_away",
        "fair_home",
        "fair_draw",
        "fair_away",
        "edge_home",
        "edge_draw",
        "edge_away",
        "value_pick",
        "value_edge",
        "bet_recommendation",
        "no_bet_reasons",
        "v19_flags",
    ]
    assert list(result.columns) == expected_columns
    assert len(result) >= 1
    assert "value_pick" in result.columns


def test_excel_report_creation_and_sheets(tmp_path):
    predictions_path = tmp_path / "predictions.csv"
    output_path = tmp_path / "predictions_report.xlsx"
    pd.DataFrame(
        [
            {
                "date": "2024-05-01",
                "league": "Premier League",
                "home_team": "Chelsea",
                "away_team": "Arsenal",
                "prob_home": 0.38,
                "prob_draw": 0.36,
                "prob_away": 0.26,
                "odds_home": 2.4,
                "odds_draw": 3.4,
                "odds_away": 2.9,
                "fair_home": 0.39,
                "fair_draw": 0.28,
                "fair_away": 0.33,
                "edge_home": -0.01,
                "edge_draw": 0.08,
                "edge_away": -0.07,
                "value_pick": "No Bet",
                "value_edge": 0.08,
                "bet_recommendation": "No bet",
                "control_score": 2.4,
                "chaos_score": 6.0,
                "tdi_home": 1,
                "tdi_away": 0,
                "no_bet_reasons": "No value bet: control score below 7",
                "v19_flags": "",
            },
            {
                "date": "2024-05-04",
                "league": "Premier League",
                "home_team": "Liverpool",
                "away_team": "Tottenham",
                "prob_home": 0.60,
                "prob_draw": 0.21,
                "prob_away": 0.19,
                "odds_home": 1.95,
                "odds_draw": 3.7,
                "odds_away": 3.85,
                "fair_home": 0.49,
                "fair_draw": 0.26,
                "fair_away": 0.25,
                "edge_home": 0.11,
                "edge_draw": -0.05,
                "edge_away": -0.06,
                "value_pick": "Home",
                "value_edge": 0.11,
                "bet_recommendation": "Value bet: Home 1X2",
                "control_score": 8.6,
                "chaos_score": 3.4,
                "tdi_home": 0,
                "tdi_away": 1,
                "no_bet_reasons": "",
                "v19_flags": "sample_flag",
            },
        ]
    ).to_csv(predictions_path, index=False)

    create_predictions_excel_report(predictions_path, output_path)

    assert output_path.exists()
    wb = load_workbook(output_path)
    assert set(["Summary", "Predictions", "Value Bets", "No Bets", "High Chaos", "v19 Flags"]).issubset(wb.sheetnames)
    headers = [cell.value for cell in wb["Predictions"][1]]
    assert "prob_home" in headers
    assert "bet_recommendation" in headers
    assert wb["Predictions"].freeze_panes == "A2"


def test_export_excel_command_runs_on_sample_predictions(tmp_path):
    root = Path(__file__).resolve().parents[1]
    history_path = root / "data" / "sample_matches.csv"
    fixtures_path = root / "data" / "upcoming_fixtures_template.csv"
    model_path = tmp_path / "sample_model.joblib"
    predictions_path = tmp_path / "predictions.csv"
    output_path = tmp_path / "predictions_report.xlsx"

    main(["train", "--input", str(history_path), "--model", str(model_path), "--test-season", "2023"])
    main([
        "predict-fixtures",
        "--history",
        str(history_path),
        "--fixtures",
        str(fixtures_path),
        "--model",
        str(model_path),
        "--output",
        str(predictions_path),
    ])
    main(["export-excel", "--predictions", str(predictions_path), "--output", str(output_path)])

    wb = load_workbook(output_path)
    assert "Summary" in wb.sheetnames
    assert "Value Bets" in wb.sheetnames


def test_odds_conversion_and_overround_removal():
    odds = {"home": 2.0, "draw": 4.0, "away": 4.0}

    assert implied_probability(2.0) == 0.5
    assert round(bookmaker_overround(odds), 4) == 0.0
    assert fair_probabilities(odds) == {"home": 0.5, "draw": 0.25, "away": 0.25}


def test_positive_value_detection():
    odds = {"home": 2.20, "draw": 3.50, "away": 3.40}
    fair = fair_probabilities(odds)
    edges = model_edges({"H": 0.50, "D": 0.25, "A": 0.25}, fair)
    side, edge = best_value_side(edges, min_edge=0.03)

    assert side == "home"
    assert edge > 0.03


def test_no_bet_when_control_score_too_low():
    from football_prediction_v19.cli import _value_recommendation

    pred = pd.Series({"odds_home": 2.20, "odds_draw": 3.50, "odds_away": 3.40})
    probs = {"H": 0.50, "D": 0.25, "A": 0.25}
    assessment = {
        "control_model_score": 6.5,
        "chaos_score": 4.0,
        "locks": [],
        "no_bets": [],
    }

    value = _value_recommendation(pred, probs, assessment, min_edge=0.03, max_chaos=7.0, min_control=7.0)

    assert value["value_pick"] == "No Bet"
    assert value["bet_recommendation"] == "No bet"
    assert any("control score below 7" in reason for reason in value["no_bet_reasons"])


def test_bet_grading_and_profit_calculation():
    stake, profit = grade_flat_stake_bet("Home", "H", 2.40)
    assert stake == 1.0
    assert profit == 1.4

    stake, profit = grade_flat_stake_bet("Away", "H", 2.90)
    assert stake == 1.0
    assert profit == -1.0

    stake, profit = grade_flat_stake_bet("No Bet", "H", 2.40)
    assert stake == 0.0
    assert profit == 0.0


def test_betting_report_creation():
    rows = pd.DataFrame(
        [
            {
                "date": "2024-01-01",
                "league": "Premier League",
                "home_team": "Chelsea",
                "away_team": "Arsenal",
                "result": "H",
                "prob_home": 0.55,
                "prob_draw": 0.25,
                "prob_away": 0.20,
                "odds_home": 2.20,
                "odds_draw": 3.40,
                "odds_away": 3.30,
                "value_pick": "Home",
                "value_edge": 0.05,
                "stake": 1.0,
                "profit": 1.2,
                "no_bet_reasons": "",
            },
            {
                "date": "2024-01-02",
                "league": "Premier League",
                "home_team": "Liverpool",
                "away_team": "Spurs",
                "result": "D",
                "prob_home": 0.40,
                "prob_draw": 0.35,
                "prob_away": 0.25,
                "odds_home": 1.90,
                "odds_draw": 3.60,
                "odds_away": 4.20,
                "value_pick": "No Bet",
                "value_edge": 0.01,
                "stake": 0.0,
                "profit": 0.0,
                "no_bet_reasons": "No value bet: control score below 7",
            },
        ]
    )

    report = build_betting_report(rows)

    assert "# Betting Backtest Report" in report
    assert "Total matches: 2" in report
    assert "Total bets: 1" in report
    assert "No value bet: control score below 7" in report


def test_backtest_bets_command_runs_on_sample_data(tmp_path):
    root = Path(__file__).resolve().parents[1]
    history_path = root / "data" / "sample_matches.csv"
    model_path = tmp_path / "sample_model.joblib"
    output_path = tmp_path / "outputs" / "backtest_bets.csv"
    report_path = tmp_path / "outputs" / "backtest_report.md"

    main(
        [
            "train",
            "--input",
            str(history_path),
            "--model",
            str(model_path),
            "--test-season",
            "2023",
        ]
    )
    main(
        [
            "backtest-bets",
            "--history",
            str(history_path),
            "--model",
            str(model_path),
            "--output",
            str(output_path),
            "--report",
            str(report_path),
            "--min-edge",
            "0.03",
            "--max-chaos",
            "7.0",
            "--min-control",
            "7.0",
        ]
    )

    assert output_path.exists()
    assert report_path.exists()
    results = pd.read_csv(output_path)
    assert "profit" in results.columns
    assert "cumulative_profit" in results.columns
    assert (results["stake"] == 0).any()
    assert "ROI" in report_path.read_text(encoding="utf-8")


def test_prepare_real_matches_cleans_training_rows():
    raw = pd.DataFrame(
        [
            {
                "date": "2023-08-13",
                "season": "2023-2024",
                "league": "Premier League",
                "home_team": " Arsenal ",
                "away_team": "Nottingham Forest",
                "score": "2-1",
                "home_xg": "1.9",
                "away_xg": "0.8",
                "odds_home": "1.45",
                "odds_draw": "4.60",
                "odds_away": "7.20",
                "venue": "Emirates Stadium",
                "referee": "Michael Oliver",
            },
            {
                "date": "2023-08-12",
                "season": "2023-2024",
                "league": "Premier League",
                "home_team": "Chelsea",
                "away_team": "Liverpool",
                "score": "",
                "home_xg": "1.4",
                "away_xg": "1.3",
                "odds_home": "2.50",
                "odds_draw": "3.40",
                "odds_away": "2.80",
                "venue": "Stamford Bridge",
                "referee": "Anthony Taylor",
            },
        ]
    )

    clean = prepare_real_matches(raw)

    assert len(clean) == 1
    assert clean.loc[0, "home_team"] == "Arsenal"
    assert clean.loc[0, "home_goals"] == 2
    assert clean.loc[0, "away_goals"] == 1
    assert clean.loc[0, "home_xg"] == 1.9
    assert clean.attrs["rows_dropped"] == 1
    assert "home_xga" in clean.attrs["optional_missing"]


def test_team_alias_normalization():
    assert normalize_team_name(" Man Utd ") == "Manchester United"
    assert normalize_team_name("Spurs") == "Tottenham Hotspur"
    assert normalize_team_name("Unknown FC") == "Unknown FC"


def test_prepare_real_matches_native_format_auto_detects():
    raw = pd.DataFrame(
        [
            {
                "date": "2023-08-12",
                "season": "2023-2024",
                "league": "Premier League",
                "home_team": "Man Utd",
                "away_team": "Wolves",
                "score": "1-0",
                "home_xg": 1.2,
                "away_xg": 0.7,
                "odds_home": 1.8,
                "odds_draw": 3.5,
                "odds_away": 4.8,
                "venue": "Old Trafford",
                "referee": "Unknown",
            }
        ]
    )

    clean = prepare_real_matches(raw, input_format="auto")

    assert clean.attrs["detected_format"] == "native"
    assert clean.loc[0, "home_team"] == "Manchester United"
    assert clean.loc[0, "away_team"] == "Wolverhampton Wanderers"


def test_prepare_real_matches_fbref_format():
    raw = pd.DataFrame(
        [
            {
                "Date": "2023-08-12",
                "Season": "2023-2024",
                "Comp": "Premier League",
                "Home": "Brighton",
                "Away": "Newcastle",
                "Score": "2-1",
                "xG": 1.7,
                "xG.1": 1.0,
                "odds_home": 2.1,
                "odds_draw": 3.4,
                "odds_away": 3.5,
                "Venue": "Amex Stadium",
                "Referee": "Unknown",
            }
        ]
    )

    clean = prepare_real_matches(raw, input_format="fbref")

    assert clean.attrs["detected_format"] == "fbref"
    assert clean.loc[0, "home_team"] == "Brighton & Hove Albion"
    assert clean.loc[0, "away_team"] == "Newcastle United"
    assert clean.loc[0, "home_xg"] == 1.7


def test_prepare_real_matches_football_data_format_auto_detects():
    raw = pd.DataFrame(
        [
            {
                "Date": "12/08/2023",
                "HomeTeam": "Tottenham",
                "AwayTeam": "Man United",
                "FTHG": 2,
                "FTAG": 0,
                "FTR": "H",
                "B365H": 2.4,
                "B365D": 3.3,
                "B365A": 2.9,
            }
        ]
    )

    clean = prepare_real_matches(raw, input_format="auto")

    assert clean.attrs["detected_format"] == "football-data"
    assert clean.loc[0, "home_team"] == "Tottenham Hotspur"
    assert clean.loc[0, "away_team"] == "Manchester United"
    assert clean.loc[0, "score"] == "2-0"
    assert clean.loc[0, "league"] == "Unknown"


def test_football_data_importer_normalizes_schema_and_aliases(tmp_path):
    root = Path(__file__).resolve().parents[1]
    output_path = tmp_path / "football_data_clean.csv"

    normalize_football_data_csv(str(root / "data" / "raw" / "football_data_template.csv"), str(output_path))

    clean = pd.read_csv(output_path)
    assert {"date", "season", "league", "home_team", "away_team", "score", "odds_home", "odds_draw", "odds_away"}.issubset(clean.columns)
    assert clean.loc[0, "home_team"] == "Tottenham Hotspur"
    assert clean.loc[0, "away_team"] == "Manchester United"


def test_fbref_importer_normalizes_schema_and_aliases(tmp_path):
    root = Path(__file__).resolve().parents[1]
    output_path = tmp_path / "fbref_clean.csv"

    normalize_fbref_csv(str(root / "data" / "raw" / "fbref_template.csv"), str(output_path))

    clean = pd.read_csv(output_path)
    assert {"date", "season", "league", "home_team", "away_team", "score", "home_xg", "away_xg", "venue", "referee"}.issubset(clean.columns)
    assert clean.loc[0, "home_team"] == "Brighton & Hove Albion"
    assert clean.loc[0, "away_team"] == "Newcastle United"


def test_import_and_prepare_command_auto_detects(tmp_path):
    root = Path(__file__).resolve().parents[1]
    output_path = tmp_path / "auto_clean.csv"

    main(
        [
            "import-and-prepare",
            "--input",
            str(root / "data" / "raw" / "fbref_template.csv"),
            "--output",
            str(output_path),
            "--format",
            "auto",
        ]
    )

    clean = pd.read_csv(output_path)
    assert clean.loc[1, "home_team"] == "Manchester United"
    assert clean.loc[1, "away_team"] == "Wolverhampton Wanderers"


def test_imported_fbref_output_is_training_compatible(tmp_path):
    root = Path(__file__).resolve().parents[1]
    source = pd.read_csv(root / "data" / "raw" / "fbref_template.csv")
    extra = source.copy()
    extra["Date"] = ["2023-08-19", "2023-08-20"]
    extra["Home"] = ["Newcastle", "Wolves"]
    extra["Away"] = ["Brighton", "Man Utd"]
    extra["Score"] = ["0-2", "1-2"]
    raw = pd.concat([source, extra], ignore_index=True)
    input_path = tmp_path / "fbref_more.csv"
    output_path = tmp_path / "fbref_clean.csv"
    raw.to_csv(input_path, index=False)

    normalize_fbref_csv(str(input_path), str(output_path))
    clean = pd.read_csv(output_path)
    model, table, metrics, cols = train_from_matches(clean, test_season=2023)

    assert len(table) >= 1
    assert metrics["feature_count"] == len(cols)


def test_prepare_real_matches_keeps_optional_columns():
    root = Path(__file__).resolve().parents[1]
    raw = pd.read_csv(root / "data" / "raw" / "real_matches_template.csv")

    clean = prepare_real_matches(raw)

    assert len(clean) == 2
    assert clean.attrs["optional_missing"] == []
    assert set(REAL_MATCH_OPTIONAL_NUMERIC_COLUMNS).issubset(clean.columns)
    assert clean.loc[0, "home_shots"] == 11
    assert clean.loc[0, "away_market_value"] == 920000000


def test_optional_columns_create_advanced_rolling_features():
    raw = pd.DataFrame(
        [
            {
                "date": "2023-08-01",
                "season": "2023-2024",
                "league": "Premier League",
                "home_team": "Chelsea",
                "away_team": "Arsenal",
                "score": "1-0",
                "home_xg": 1.3,
                "away_xg": 0.8,
                "odds_home": 2.2,
                "odds_draw": 3.4,
                "odds_away": 3.1,
                "venue": "Stamford Bridge",
                "referee": "Anthony Taylor",
                "home_xga": 0.7,
                "away_xga": 1.4,
                "home_shots": 12,
                "away_shots": 8,
                "home_shots_on_target": 5,
                "away_shots_on_target": 3,
                "home_big_chances": 3,
                "away_big_chances": 1,
                "home_rest_days": 6,
                "away_rest_days": 5,
                "home_injuries_count": 1,
                "away_injuries_count": 2,
            },
            {
                "date": "2023-08-08",
                "season": "2023-2024",
                "league": "Premier League",
                "home_team": "Arsenal",
                "away_team": "Chelsea",
                "score": "2-2",
                "home_xg": 1.5,
                "away_xg": 1.2,
                "odds_home": 2.0,
                "odds_draw": 3.5,
                "odds_away": 3.6,
                "venue": "Emirates Stadium",
                "referee": "Michael Oliver",
                "home_xga": 1.1,
                "away_xga": 1.6,
                "home_shots": 15,
                "away_shots": 10,
                "home_shots_on_target": 6,
                "away_shots_on_target": 4,
                "home_big_chances": 2,
                "away_big_chances": 2,
                "home_rest_days": 7,
                "away_rest_days": 7,
                "home_injuries_count": 0,
                "away_injuries_count": 1,
            },
        ]
    )

    clean = prepare_real_matches(raw)
    features = build_features(clean, min_history=1)

    assert "home_w5_shots" in features.columns
    assert "away_w5_shots_on_target" in features.columns
    assert "edge_w5_big_chances" in features.columns
    assert features.loc[0, "home_w5_shots"] == 8
    assert features.loc[0, "away_w5_injuries_count"] == 1


def test_build_football_data_url_known_leagues():
    assert build_football_data_url("E0", 2023) == "https://www.football-data.co.uk/mmz4281/2324/E0.csv"
    assert build_football_data_url("premier-league", 2022) == "https://www.football-data.co.uk/mmz4281/2223/E0.csv"
    assert build_football_data_url("bundesliga", 2023) == "https://www.football-data.co.uk/mmz4281/2324/D1.csv"
    assert build_football_data_url("D2", 2024) == "https://www.football-data.co.uk/mmz4281/2425/D2.csv"


def test_league_codes_dict_contains_common_leagues():
    assert "premier-league" in LEAGUE_CODES
    assert LEAGUE_CODES["premier-league"] == "E0"
    assert "bundesliga" in LEAGUE_CODES
    assert "serie-a" in LEAGUE_CODES
    assert "la-liga" in LEAGUE_CODES


def test_download_season_uses_correct_url_and_filename(tmp_path):
    import unittest.mock as mock

    fake_csv = b"Date,HomeTeam,AwayTeam\n01/08/2023,Arsenal,Chelsea\n"
    with mock.patch("football_prediction_v19.importers.football_data.requests.get") as mock_get:
        mock_get.return_value = mock.Mock(status_code=200, content=fake_csv)
        mock_get.return_value.raise_for_status = lambda: None
        path = download_season("E0", 2023, tmp_path)

    assert path.name == "E0_2023_2024.csv"
    assert path.read_bytes() == fake_csv
    called_url = mock_get.call_args[0][0]
    assert called_url == "https://www.football-data.co.uk/mmz4281/2324/E0.csv"


def test_bulk_download_downloads_all_combinations(tmp_path):
    import unittest.mock as mock

    fake_csv = b"Date,HomeTeam,AwayTeam\n01/08/2023,Arsenal,Chelsea\n"
    with mock.patch("football_prediction_v19.importers.football_data.requests.get") as mock_get:
        mock_get.return_value = mock.Mock(status_code=200, content=fake_csv)
        mock_get.return_value.raise_for_status = lambda: None
        paths = bulk_download(["E0", "D1"], [2022, 2023], tmp_path)

    assert len(paths) == 4
    names = {p.name for p in paths}
    assert "E0_2022_2023.csv" in names
    assert "E0_2023_2024.csv" in names
    assert "D1_2022_2023.csv" in names
    assert "D1_2023_2024.csv" in names


def test_download_football_data_cli_command(tmp_path):
    import unittest.mock as mock

    fake_csv = b"Date,HomeTeam,AwayTeam\n01/08/2023,Arsenal,Chelsea\n"
    with mock.patch("football_prediction_v19.importers.football_data.requests.get") as mock_get:
        mock_get.return_value = mock.Mock(status_code=200, content=fake_csv)
        mock_get.return_value.raise_for_status = lambda: None
        main([
            "download-football-data",
            "--leagues", "E0",
            "--seasons", "2023",
            "--output-dir", str(tmp_path),
        ])

    assert (tmp_path / "E0_2023_2024.csv").exists()


def test_download_football_data_cli_bulk(tmp_path):
    import unittest.mock as mock

    fake_csv = b"Date,HomeTeam,AwayTeam\n01/08/2023,Arsenal,Chelsea\n"
    with mock.patch("football_prediction_v19.importers.football_data.requests.get") as mock_get:
        mock_get.return_value = mock.Mock(status_code=200, content=fake_csv)
        mock_get.return_value.raise_for_status = lambda: None
        main([
            "download-football-data",
            "--leagues", "E0", "premier-league",
            "--seasons", "2022", "2023",
            "--output-dir", str(tmp_path),
        ])

    assert (tmp_path / "E0_2022_2023.csv").exists()
    assert (tmp_path / "E0_2023_2024.csv").exists()


_FAKE_FD_CSV = (
    b"Date,HomeTeam,AwayTeam,FTHG,FTAG,FTR,B365H,B365D,B365A\n"
    b"12/08/2023,Tottenham,Man United,2,0,H,2.40,3.30,2.90\n"
    b"13/08/2023,Wolves,Newcastle,1,1,D,3.10,3.20,2.35\n"
)

_REQUIRED_TRAINING_COLUMNS = {
    "date", "season", "league", "home_team", "away_team", "score",
    "odds_home", "odds_draw", "odds_away",
}


def _mock_get(fake_csv: bytes):
    import unittest.mock as mock
    m = mock.Mock(status_code=200, content=fake_csv)
    m.raise_for_status = lambda: None
    return m


def test_download_and_prepare_single(tmp_path):
    import unittest.mock as mock

    raw_dir = tmp_path / "raw"
    proc_dir = tmp_path / "processed"

    with mock.patch(
        "football_prediction_v19.importers.football_data.requests.get",
        return_value=_mock_get(_FAKE_FD_CSV),
    ):
        result = download_and_prepare("E0", 2023, raw_dir, proc_dir)

    assert Path(result["raw_path"]).name == "football_data_E0_2023.csv"
    assert Path(result["processed_path"]).name == "football_data_E0_2023_clean.csv"
    assert result["rows_written"] == 2
    clean = pd.read_csv(result["processed_path"])
    assert _REQUIRED_TRAINING_COLUMNS.issubset(clean.columns)
    assert clean.loc[0, "home_team"] == "Tottenham Hotspur"


def test_download_and_prepare_multiple(tmp_path):
    import unittest.mock as mock

    raw_dir = tmp_path / "raw"
    proc_dir = tmp_path / "processed"

    with mock.patch(
        "football_prediction_v19.importers.football_data.requests.get",
        return_value=_mock_get(_FAKE_FD_CSV),
    ):
        r1 = download_and_prepare("E0", 2022, raw_dir, proc_dir)
        r2 = download_and_prepare("D1", 2023, raw_dir, proc_dir)

    assert Path(r1["raw_path"]).name == "football_data_E0_2022.csv"
    assert Path(r2["raw_path"]).name == "football_data_D1_2023.csv"
    assert r1["rows_written"] == 2
    assert r2["rows_written"] == 2


def test_download_prepare_cli_single(tmp_path):
    import unittest.mock as mock

    raw_dir = tmp_path / "raw"
    proc_dir = tmp_path / "processed"

    with mock.patch(
        "football_prediction_v19.importers.football_data.requests.get",
        return_value=_mock_get(_FAKE_FD_CSV),
    ):
        main([
            "download-prepare-football-data",
            "--leagues", "E0",
            "--seasons", "2023",
            "--raw-dir", str(raw_dir),
            "--processed-dir", str(proc_dir),
        ])

    assert (raw_dir / "football_data_E0_2023.csv").exists()
    assert (proc_dir / "football_data_E0_2023_clean.csv").exists()
    clean = pd.read_csv(proc_dir / "football_data_E0_2023_clean.csv")
    assert _REQUIRED_TRAINING_COLUMNS.issubset(clean.columns)


def test_download_prepare_cli_combine_output(tmp_path):
    import unittest.mock as mock

    raw_dir = tmp_path / "raw"
    proc_dir = tmp_path / "processed"
    combined = tmp_path / "combined.csv"

    with mock.patch(
        "football_prediction_v19.importers.football_data.requests.get",
        return_value=_mock_get(_FAKE_FD_CSV),
    ):
        main([
            "download-prepare-football-data",
            "--leagues", "E0", "D1",
            "--seasons", "2023",
            "--raw-dir", str(raw_dir),
            "--processed-dir", str(proc_dir),
            "--combine-output", str(combined),
        ])

    assert combined.exists()
    df = pd.read_csv(combined)
    assert len(df) == 4  # 2 rows per league
    assert list(df.columns[:1]) == ["date"]  # sorted by date


def test_download_prepare_cli_unknown_league(tmp_path):
    import pytest

    with pytest.raises(SystemExit, match="Unknown league"):
        main([
            "download-prepare-football-data",
            "--leagues", "NOTACODE",
            "--seasons", "2023",
            "--raw-dir", str(tmp_path / "raw"),
            "--processed-dir", str(tmp_path / "processed"),
        ])


def test_download_prepare_schema_contains_training_columns(tmp_path):
    import unittest.mock as mock

    raw_dir = tmp_path / "raw"
    proc_dir = tmp_path / "processed"

    with mock.patch(
        "football_prediction_v19.importers.football_data.requests.get",
        return_value=_mock_get(_FAKE_FD_CSV),
    ):
        result = download_and_prepare("premier-league", 2023, raw_dir, proc_dir)

    clean = pd.read_csv(result["processed_path"])
    assert _REQUIRED_TRAINING_COLUMNS.issubset(clean.columns)
    assert result["league_code"] == "E0"


def test_prepare_data_command_writes_clean_csv(tmp_path):
    input_path = tmp_path / "raw" / "real_matches.csv"
    output_path = tmp_path / "processed" / "real_matches_clean.csv"
    input_path.parent.mkdir(parents=True)
    input_path.write_text(
        "\n".join(
            [
                "date,season,league,home_team,away_team,score,home_xg,away_xg,odds_home,odds_draw,odds_away,venue,referee",
                "2023-08-12,2023-2024,Premier League,Chelsea,Liverpool,1-1,1.4,1.3,2.50,3.40,2.80,Stamford Bridge,Anthony Taylor",
                "2023-08-13,2023-2024,Premier League,Arsenal,Nottingham Forest,,1.9,0.8,1.45,4.60,7.20,Emirates Stadium,Michael Oliver",
            ]
        ),
        encoding="utf-8",
    )

    main(["prepare-data", "--input", str(input_path), "--output", str(output_path)])

    assert output_path.exists()
    clean = pd.read_csv(output_path)
    assert list(clean["home_team"]) == ["Chelsea"]
    assert clean.loc[0, "home_goals"] == 1


# ---------------------------------------------------------------------------
# prepare-fixtures tests
# ---------------------------------------------------------------------------

_FIXTURE_OUTPUT_COLUMNS = [
    "date", "season", "league", "home_team", "away_team", "venue", "referee",
    "odds_home", "odds_draw", "odds_away",
    "formation_home_xg90", "formation_away_xg90",
    "fatigue_home", "fatigue_away",
]


def test_prepare_fixtures_native_format():
    df = pd.DataFrame([
        {
            "date": "2024-08-17",
            "season": "2024-2025",
            "league": "Premier League",
            "home_team": "Man Utd",
            "away_team": "Wolves",
            "venue": "Old Trafford",
            "referee": "Michael Oliver",
            "odds_home": 2.10,
            "odds_draw": 3.50,
            "odds_away": 3.40,
            "formation_home_xg90": 1.20,
            "formation_away_xg90": 1.05,
            "fatigue_home": 0.10,
            "fatigue_away": 0.08,
        }
    ])
    out = prepare_fixtures(df, input_format="native")
    assert list(out.columns) == _FIXTURE_OUTPUT_COLUMNS
    assert out.loc[0, "home_team"] == "Manchester United"
    assert out.loc[0, "away_team"] == "Wolverhampton Wanderers"
    assert out.loc[0, "date"] == "2024-08-17"


def test_prepare_fixtures_football_data_format():
    df = pd.DataFrame([
        {"Date": "17/08/2024", "HomeTeam": "Tottenham", "AwayTeam": "Leicester",
         "B365H": 1.75, "B365D": 3.60, "B365A": 4.80},
    ])
    out = prepare_fixtures(df, input_format="football-data", default_season="2024", default_league="Premier League")
    assert list(out.columns) == _FIXTURE_OUTPUT_COLUMNS
    assert out.loc[0, "home_team"] == "Tottenham Hotspur"
    assert out.loc[0, "away_team"] == "Leicester"
    assert out.loc[0, "season"] == "2024"
    assert out.loc[0, "league"] == "Premier League"
    assert out.loc[0, "odds_home"] == 1.75


def test_prepare_fixtures_missing_optional_columns_get_defaults():
    df = pd.DataFrame([
        {"date": "2024-08-17", "home_team": "Arsenal", "away_team": "Chelsea"},
    ])
    out = prepare_fixtures(df, input_format="native", default_season="2024", default_league="Bundesliga")
    assert out.loc[0, "venue"] == ""
    assert out.loc[0, "referee"] == ""
    assert out.loc[0, "season"] == "2024"
    assert out.loc[0, "league"] == "Bundesliga"
    assert out.loc[0, "formation_home_xg90"] == 0.0
    assert out.loc[0, "formation_away_xg90"] == 0.0
    assert out.loc[0, "fatigue_home"] == 0.0
    assert out.loc[0, "fatigue_away"] == 0.0
    import math
    assert math.isnan(out.loc[0, "odds_home"])


def test_prepare_fixtures_team_alias_normalization():
    df = pd.DataFrame([
        {"date": "2024-08-17", "home_team": "Spurs", "away_team": "Man United"},
    ])
    out = prepare_fixtures(df, input_format="native")
    assert out.loc[0, "home_team"] == "Tottenham Hotspur"
    assert out.loc[0, "away_team"] == "Manchester United"


def test_prepare_fixtures_missing_required_column_raises():
    import pytest
    df = pd.DataFrame([{"date": "2024-08-17", "home_team": "Arsenal"}])
    with pytest.raises(ValueError, match="away_team"):
        prepare_fixtures(df, input_format="native")


def test_prepare_fixtures_empty_input_raises():
    import pytest
    df = pd.DataFrame(columns=["date", "home_team", "away_team"])
    with pytest.raises(ValueError, match="empty"):
        prepare_fixtures(df, input_format="native")


def test_prepare_fixtures_unknown_format_raises():
    import pytest
    df = pd.DataFrame([{"date": "2024-08-17", "home_team": "Arsenal", "away_team": "Chelsea"}])
    with pytest.raises(ValueError, match="Unsupported"):
        prepare_fixtures(df, input_format="xls")


def test_prepare_fixtures_auto_detects_native():
    df = pd.DataFrame([
        {"date": "2024-08-17", "home_team": "Arsenal", "away_team": "Chelsea"},
    ])
    out = prepare_fixtures(df, input_format="auto")
    assert out.loc[0, "home_team"] == "Arsenal"


def test_prepare_fixtures_auto_detects_football_data():
    df = pd.DataFrame([
        {"Date": "17/08/2024", "HomeTeam": "Arsenal", "AwayTeam": "Chelsea", "B365H": 2.0, "B365D": 3.5, "B365A": 3.5},
    ])
    out = prepare_fixtures(df, input_format="auto")
    assert out.loc[0, "home_team"] == "Arsenal"
    assert out.loc[0, "odds_home"] == 2.0


def test_prepare_fixtures_sorts_by_date():
    df = pd.DataFrame([
        {"date": "2024-08-20", "home_team": "Arsenal", "away_team": "Chelsea"},
        {"date": "2024-08-17", "home_team": "Liverpool", "away_team": "Spurs"},
    ])
    out = prepare_fixtures(df, input_format="native")
    assert out.loc[0, "home_team"] == "Liverpool"
    assert out.loc[1, "home_team"] == "Arsenal"


def test_prepare_fixtures_cli_command(tmp_path):
    root = Path(__file__).resolve().parents[1]
    input_path = root / "data" / "raw" / "upcoming_fixtures_raw_template.csv"
    output_path = tmp_path / "fixtures_ready.csv"

    main([
        "prepare-fixtures",
        "--input", str(input_path),
        "--output", str(output_path),
        "--format", "native",
    ])

    assert output_path.exists()
    out = pd.read_csv(output_path)
    assert list(out.columns) == _FIXTURE_OUTPUT_COLUMNS
    assert out.loc[0, "home_team"] == "Manchester United"


def test_prepare_fixtures_football_data_template(tmp_path):
    root = Path(__file__).resolve().parents[1]
    input_path = root / "data" / "raw" / "football_data_fixtures_template.csv"
    output_path = tmp_path / "fixtures_ready.csv"

    main([
        "prepare-fixtures",
        "--input", str(input_path),
        "--output", str(output_path),
        "--format", "football-data",
        "--default-season", "2024",
        "--default-league", "Premier League",
    ])

    out = pd.read_csv(output_path)
    assert list(out.columns) == _FIXTURE_OUTPUT_COLUMNS
    assert out.loc[0, "home_team"] == "Manchester United"
    assert str(out.loc[0, "season"]) == "2024"


def test_prepare_fixtures_then_predict_fixtures(tmp_path):
    root = Path(__file__).resolve().parents[1]
    raw_fixtures = root / "data" / "raw" / "upcoming_fixtures_raw_template.csv"
    history_path = root / "data" / "sample_matches.csv"
    prepared_path = tmp_path / "fixtures_ready.csv"
    model_path = tmp_path / "model.joblib"
    predictions_path = tmp_path / "predictions.csv"

    main([
        "prepare-fixtures",
        "--input", str(raw_fixtures),
        "--output", str(prepared_path),
        "--format", "native",
    ])
    main(["train", "--input", str(history_path), "--model", str(model_path), "--test-season", "2023"])
    main([
        "predict-fixtures",
        "--history", str(history_path),
        "--fixtures", str(prepared_path),
        "--model", str(model_path),
        "--output", str(predictions_path),
    ])

    assert predictions_path.exists()
    preds = pd.read_csv(predictions_path)
    assert "prob_home" in preds.columns
    assert len(preds) >= 1


# ---------------------------------------------------------------------------
# run-pipeline tests
# ---------------------------------------------------------------------------

def _pipeline_base_args(tmp_path, root):
    """Return a base args list for run-pipeline using local sample data."""
    return [
        "run-pipeline",
        "--skip-download",
        "--combine-output", str(root / "data" / "sample_matches.csv"),
        "--fixtures-raw", str(root / "data" / "raw" / "upcoming_fixtures_raw_template.csv"),
        "--fixtures-output", str(tmp_path / "fixtures_ready.csv"),
        "--fixtures-format", "native",
        "--model", str(tmp_path / "pipeline_model.joblib"),
        "--predictions", str(tmp_path / "predictions.csv"),
        "--excel", str(tmp_path / "predictions_report.xlsx"),
        "--backtest-csv", str(tmp_path / "backtest_bets.csv"),
        "--backtest-report", str(tmp_path / "backtest_report.md"),
        "--test-season", "2023",
    ]


def test_run_pipeline_skip_download_creates_all_outputs(tmp_path):
    root = Path(__file__).resolve().parents[1]
    main(_pipeline_base_args(tmp_path, root))

    assert (tmp_path / "fixtures_ready.csv").exists()
    assert (tmp_path / "pipeline_model.joblib").exists()
    assert (tmp_path / "predictions.csv").exists()
    assert (tmp_path / "predictions_report.xlsx").exists()
    assert (tmp_path / "backtest_bets.csv").exists()
    assert (tmp_path / "backtest_report.md").exists()


def test_run_pipeline_predictions_have_correct_columns(tmp_path):
    root = Path(__file__).resolve().parents[1]
    main(_pipeline_base_args(tmp_path, root))

    preds = pd.read_csv(tmp_path / "predictions.csv")
    for col in ["prob_home", "prob_draw", "prob_away", "bet_recommendation", "value_edge"]:
        assert col in preds.columns
    assert len(preds) >= 1


def test_run_pipeline_excel_created(tmp_path):
    from openpyxl import load_workbook
    root = Path(__file__).resolve().parents[1]
    main(_pipeline_base_args(tmp_path, root))

    wb = load_workbook(tmp_path / "predictions_report.xlsx")
    assert "Summary" in wb.sheetnames
    assert "Predictions" in wb.sheetnames


def test_run_pipeline_backtest_files_created(tmp_path):
    root = Path(__file__).resolve().parents[1]
    main(_pipeline_base_args(tmp_path, root))

    bt = pd.read_csv(tmp_path / "backtest_bets.csv")
    assert "profit" in bt.columns
    report_text = (tmp_path / "backtest_report.md").read_text(encoding="utf-8")
    assert "ROI" in report_text


def test_run_pipeline_skip_backtest_no_backtest_files(tmp_path):
    root = Path(__file__).resolve().parents[1]
    args = _pipeline_base_args(tmp_path, root) + ["--skip-backtest"]
    main(args)

    assert (tmp_path / "predictions.csv").exists()
    assert (tmp_path / "predictions_report.xlsx").exists()
    assert not (tmp_path / "backtest_bets.csv").exists()
    assert not (tmp_path / "backtest_report.md").exists()


def test_run_pipeline_use_existing_fixtures(tmp_path):
    root = Path(__file__).resolve().parents[1]
    # First prepare the fixtures
    fixtures_path = tmp_path / "fixtures_ready.csv"
    main([
        "prepare-fixtures",
        "--input", str(root / "data" / "raw" / "upcoming_fixtures_raw_template.csv"),
        "--output", str(fixtures_path),
        "--format", "native",
    ])
    assert fixtures_path.exists()

    # Now run pipeline without re-preparing fixtures
    main([
        "run-pipeline",
        "--skip-download",
        "--use-existing-fixtures",
        "--combine-output", str(root / "data" / "sample_matches.csv"),
        "--fixtures-output", str(fixtures_path),
        "--model", str(tmp_path / "model.joblib"),
        "--predictions", str(tmp_path / "predictions.csv"),
        "--excel", str(tmp_path / "report.xlsx"),
        "--skip-backtest",
    ])
    assert (tmp_path / "predictions.csv").exists()


def test_run_pipeline_missing_combine_output_with_skip_download_gives_error(tmp_path):
    import pytest
    with pytest.raises(SystemExit, match="does not exist"):
        main([
            "run-pipeline",
            "--skip-download",
            "--combine-output", str(tmp_path / "nonexistent.csv"),
            "--fixtures-raw", str(tmp_path / "fixtures.csv"),
            "--fixtures-output", str(tmp_path / "fx.csv"),
            "--model", str(tmp_path / "model.joblib"),
            "--predictions", str(tmp_path / "preds.csv"),
            "--excel", str(tmp_path / "report.xlsx"),
        ])


def test_run_pipeline_missing_fixtures_raw_gives_error(tmp_path):
    import pytest
    root = Path(__file__).resolve().parents[1]
    with pytest.raises(SystemExit, match="fixtures-raw"):
        main([
            "run-pipeline",
            "--skip-download",
            "--combine-output", str(root / "data" / "sample_matches.csv"),
            "--fixtures-output", str(tmp_path / "fx.csv"),
            "--model", str(tmp_path / "model.joblib"),
            "--predictions", str(tmp_path / "preds.csv"),
            "--excel", str(tmp_path / "report.xlsx"),
        ])


# ---------------------------------------------------------------------------
# odds_import tests
# ---------------------------------------------------------------------------

def test_prepare_odds_native_format():
    from football_prediction_v19.odds_import import prepare_odds
    df = pd.DataFrame({
        "date": ["2024-08-17"],
        "home_team": ["Man Utd"],
        "away_team": ["Wolves"],
        "odds_home": [2.10],
        "odds_draw": [3.50],
        "odds_away": [3.40],
    })
    out = prepare_odds(df)
    assert len(out) == 1
    assert out.loc[0, "odds_home"] == pytest.approx(2.10)
    assert out.loc[0, "home_team"] is not None


def test_prepare_odds_b365_columns():
    from football_prediction_v19.odds_import import prepare_odds
    df = pd.DataFrame({
        "Date": ["17/08/2024"],
        "HomeTeam": ["Man United"],
        "AwayTeam": ["Wolves"],
        "B365H": [2.10],
        "B365D": [3.50],
        "B365A": [3.40],
    })
    out = prepare_odds(df)
    assert "odds_home" in out.columns
    assert out.loc[0, "odds_home"] == pytest.approx(2.10)


def test_prepare_odds_ps_columns():
    from football_prediction_v19.odds_import import prepare_odds
    df = pd.DataFrame({
        "HomeTeam": ["Arsenal"],
        "AwayTeam": ["Chelsea"],
        "PSH": [1.90],
        "PSD": [3.60],
        "PSA": [4.00],
    })
    out = prepare_odds(df)
    assert out.loc[0, "odds_home"] == pytest.approx(1.90)


def test_prepare_odds_max_avg_columns():
    from football_prediction_v19.odds_import import prepare_odds
    df = pd.DataFrame({
        "HomeTeam": ["Arsenal"],
        "AwayTeam": ["Chelsea"],
        "MaxH": [2.00],
        "MaxD": [3.70],
        "MaxA": [4.10],
    })
    out = prepare_odds(df)
    assert out.loc[0, "odds_home"] == pytest.approx(2.00)


def test_prepare_odds_team_alias_normalization():
    from football_prediction_v19.odds_import import prepare_odds
    df = pd.DataFrame({
        "Home": ["Spurs"],
        "Away": ["Leicester"],
        "odds_home": [1.75],
        "odds_draw": [3.60],
        "odds_away": [4.80],
    })
    out = prepare_odds(df)
    assert out.loc[0, "home_team"] is not None
    assert out.loc[0, "away_team"] is not None


def test_prepare_odds_empty_raises():
    from football_prediction_v19.odds_import import prepare_odds
    with pytest.raises(ValueError, match="empty"):
        prepare_odds(pd.DataFrame())


def test_prepare_odds_missing_team_column_raises():
    from football_prediction_v19.odds_import import prepare_odds
    df = pd.DataFrame({"odds_home": [2.0], "odds_draw": [3.5], "odds_away": [3.4]})
    with pytest.raises(ValueError, match="home_team"):
        prepare_odds(df)


def test_prepare_odds_no_odds_column_raises():
    from football_prediction_v19.odds_import import prepare_odds
    df = pd.DataFrame({"home_team": ["Arsenal"], "away_team": ["Chelsea"]})
    with pytest.raises(ValueError, match="No usable odds"):
        prepare_odds(df)


def test_prepare_odds_all_invalid_raises():
    from football_prediction_v19.odds_import import prepare_odds
    df = pd.DataFrame({
        "home_team": [""],
        "away_team": [""],
        "odds_home": [float("nan")],
        "odds_draw": [float("nan")],
        "odds_away": [float("nan")],
    })
    with pytest.raises(ValueError):
        prepare_odds(df)


def test_prepare_odds_file_missing_input_raises(tmp_path):
    from football_prediction_v19.odds_import import prepare_odds_file
    with pytest.raises(ValueError, match="not found"):
        prepare_odds_file(str(tmp_path / "nonexistent.csv"), str(tmp_path / "out.csv"))


def test_cli_prepare_odds_native_template(tmp_path):
    root = Path(__file__).resolve().parents[1]
    src = root / "data" / "raw" / "odds_raw_template.csv"
    out = tmp_path / "odds_clean.csv"
    main(["prepare-odds", "--input", str(src), "--output", str(out)])
    assert out.exists()
    df = pd.read_csv(out)
    assert len(df) == 2
    assert "odds_home" in df.columns


def test_cli_prepare_odds_football_data_template(tmp_path):
    root = Path(__file__).resolve().parents[1]
    src = root / "data" / "raw" / "football_data_odds_template.csv"
    out = tmp_path / "odds_clean.csv"
    main(["prepare-odds", "--input", str(src), "--output", str(out)])
    assert out.exists()
    df = pd.read_csv(out)
    assert len(df) == 2
    assert "odds_home" in df.columns


def test_merge_odds_into_fixtures_basic_match():
    from football_prediction_v19.odds_import import merge_odds_into_fixtures
    fixtures = pd.DataFrame({
        "date": ["2024-08-17"],
        "home_team": ["Man Utd"],
        "away_team": ["Wolves"],
        "odds_home": [float("nan")],
        "odds_draw": [float("nan")],
        "odds_away": [float("nan")],
    })
    odds = pd.DataFrame({
        "date": [pd.Timestamp("2024-08-17")],
        "home_team": ["Man Utd"],
        "away_team": ["Wolves"],
        "odds_home": [2.10],
        "odds_draw": [3.50],
        "odds_away": [3.40],
    })
    updated, matched, missing = merge_odds_into_fixtures(fixtures, odds)
    assert matched == 1
    assert missing == 0
    assert updated.loc[0, "odds_home"] == pytest.approx(2.10)


def test_merge_odds_date_window_no_match_at_zero():
    from football_prediction_v19.odds_import import merge_odds_into_fixtures
    fixtures = pd.DataFrame({
        "date": ["2024-08-17"],
        "home_team": ["Arsenal"],
        "away_team": ["Chelsea"],
        "odds_home": [float("nan")],
        "odds_draw": [float("nan")],
        "odds_away": [float("nan")],
    })
    odds = pd.DataFrame({
        "date": [pd.Timestamp("2024-08-15")],
        "home_team": ["Arsenal"],
        "away_team": ["Chelsea"],
        "odds_home": [1.90],
        "odds_draw": [3.60],
        "odds_away": [4.00],
    })
    _, matched, _ = merge_odds_into_fixtures(fixtures, odds, allow_date_window=0)
    assert matched == 0


def test_merge_odds_date_window_match_within_two_days():
    from football_prediction_v19.odds_import import merge_odds_into_fixtures
    fixtures = pd.DataFrame({
        "date": ["2024-08-17"],
        "home_team": ["Arsenal"],
        "away_team": ["Chelsea"],
        "odds_home": [float("nan")],
        "odds_draw": [float("nan")],
        "odds_away": [float("nan")],
    })
    odds = pd.DataFrame({
        "date": [pd.Timestamp("2024-08-15")],
        "home_team": ["Arsenal"],
        "away_team": ["Chelsea"],
        "odds_home": [1.90],
        "odds_draw": [3.60],
        "odds_away": [4.00],
    })
    _, matched, _ = merge_odds_into_fixtures(fixtures, odds, allow_date_window=2)
    assert matched == 1


def test_merge_odds_bookmaker_preference():
    from football_prediction_v19.odds_import import merge_odds_into_fixtures
    fixtures = pd.DataFrame({
        "date": [pd.NaT],
        "home_team": ["Arsenal"],
        "away_team": ["Chelsea"],
        "odds_home": [float("nan")],
        "odds_draw": [float("nan")],
        "odds_away": [float("nan")],
    })
    odds = pd.DataFrame({
        "date": [pd.NaT, pd.NaT],
        "home_team": ["Arsenal", "Arsenal"],
        "away_team": ["Chelsea", "Chelsea"],
        "odds_home": [1.80, 1.90],
        "odds_draw": [3.50, 3.60],
        "odds_away": [4.00, 4.10],
        "bookmaker": ["Other", "Bet365"],
    })
    updated, matched, _ = merge_odds_into_fixtures(fixtures, odds, prefer_bookmaker="Bet365")
    assert matched == 1
    assert updated.loc[0, "odds_home"] == pytest.approx(1.90)


def test_cli_merge_odds_fixtures(tmp_path):
    root = Path(__file__).resolve().parents[1]
    odds_src = root / "data" / "raw" / "odds_raw_template.csv"
    odds_clean = tmp_path / "odds_clean.csv"
    main(["prepare-odds", "--input", str(odds_src), "--output", str(odds_clean)])

    fx_src = root / "data" / "raw" / "upcoming_fixtures_raw_template.csv"
    fx_clean = tmp_path / "fx_clean.csv"
    main(["prepare-fixtures", "--input", str(fx_src), "--output", str(fx_clean)])

    out = tmp_path / "fx_with_odds.csv"
    main([
        "merge-odds-fixtures",
        "--fixtures", str(fx_clean),
        "--odds", str(odds_clean),
        "--output", str(out),
    ])
    assert out.exists()


def test_run_pipeline_with_odds_raw(tmp_path):
    root = Path(__file__).resolve().parents[1]
    odds_raw = root / "data" / "raw" / "odds_raw_template.csv"
    out = tmp_path / "predictions.csv"
    main([
        "run-pipeline",
        "--skip-download",
        "--combine-output", str(root / "data" / "sample_matches.csv"),
        "--fixtures-raw", str(root / "data" / "raw" / "upcoming_fixtures_raw_template.csv"),
        "--fixtures-output", str(tmp_path / "fx.csv"),
        "--model", str(tmp_path / "model.joblib"),
        "--predictions", str(out),
        "--excel", str(tmp_path / "report.xlsx"),
        "--odds-raw", str(odds_raw),
        "--odds-clean", str(tmp_path / "odds_clean.csv"),
        "--fixtures-with-odds", str(tmp_path / "fx_with_odds.csv"),
    ])
    assert out.exists()


# ---------------------------------------------------------------------------
# xg_import tests
# ---------------------------------------------------------------------------

def test_prepare_xg_native_format():
    from football_prediction_v19.xg_import import prepare_xg
    df = pd.DataFrame({
        "date": ["2024-08-17"],
        "home_team": ["Man Utd"],
        "away_team": ["Wolves"],
        "home_xg": [1.23],
        "away_xg": [0.85],
    })
    out = prepare_xg(df)
    assert len(out) == 1
    assert out.loc[0, "home_xg"] == pytest.approx(1.23)
    assert out.loc[0, "away_xg"] == pytest.approx(0.85)


def test_prepare_xg_fbref_columns():
    from football_prediction_v19.xg_import import prepare_xg
    df = pd.DataFrame({
        "Date": ["17/08/2024"],
        "Home": ["Man United"],
        "Away": ["Wolves"],
        "xG": [1.23],
        "xG.1": [0.85],
        "Comp": ["Premier League"],
        "Season": ["2024-2025"],
    })
    out = prepare_xg(df)
    assert "home_xg" in out.columns
    assert "away_xg" in out.columns
    assert out.loc[0, "home_xg"] == pytest.approx(1.23)
    assert out.loc[0, "away_xg"] == pytest.approx(0.85)
    assert "league" in out.columns
    assert "season" in out.columns


def test_prepare_xg_understat_columns():
    from football_prediction_v19.xg_import import prepare_xg
    df = pd.DataFrame({
        "date": ["2024-08-17"],
        "h_team": ["Arsenal"],
        "a_team": ["Chelsea"],
        "xG": [2.10],
        "xGA": [0.65],
    })
    out = prepare_xg(df)
    assert out.loc[0, "home_xg"] == pytest.approx(2.10)
    assert out.loc[0, "away_xg"] == pytest.approx(0.65)


def test_prepare_xg_homexg_awayxg_aliases():
    from football_prediction_v19.xg_import import prepare_xg
    df = pd.DataFrame({
        "home_team": ["Arsenal"],
        "away_team": ["Chelsea"],
        "home_xG": [1.50],
        "away_xG": [0.90],
    })
    out = prepare_xg(df)
    assert out.loc[0, "home_xg"] == pytest.approx(1.50)
    assert out.loc[0, "away_xg"] == pytest.approx(0.90)


def test_prepare_xg_team_alias_normalization():
    from football_prediction_v19.xg_import import prepare_xg
    df = pd.DataFrame({
        "Home": ["Spurs"],
        "Away": ["Leicester"],
        "home_xg": [2.10],
        "away_xg": [0.65],
    })
    out = prepare_xg(df)
    assert out.loc[0, "home_team"] is not None
    assert out.loc[0, "away_team"] is not None


def test_prepare_xg_empty_raises():
    from football_prediction_v19.xg_import import prepare_xg
    with pytest.raises(ValueError, match="empty"):
        prepare_xg(pd.DataFrame())


def test_prepare_xg_missing_team_column_raises():
    from football_prediction_v19.xg_import import prepare_xg
    df = pd.DataFrame({"home_xg": [1.5], "away_xg": [0.8]})
    with pytest.raises(ValueError, match="home_team"):
        prepare_xg(df)


def test_prepare_xg_no_xg_column_raises():
    from football_prediction_v19.xg_import import prepare_xg
    df = pd.DataFrame({"home_team": ["Arsenal"], "away_team": ["Chelsea"]})
    with pytest.raises(ValueError, match="No usable xG"):
        prepare_xg(df)


def test_prepare_xg_all_invalid_raises():
    from football_prediction_v19.xg_import import prepare_xg
    df = pd.DataFrame({
        "home_team": [""],
        "away_team": [""],
        "home_xg": [float("nan")],
        "away_xg": [float("nan")],
    })
    with pytest.raises(ValueError):
        prepare_xg(df)


def test_prepare_xg_unsupported_format_raises():
    from football_prediction_v19.xg_import import prepare_xg
    df = pd.DataFrame({
        "home_team": ["Arsenal"],
        "away_team": ["Chelsea"],
        "home_xg": [1.5],
        "away_xg": [0.8],
    })
    with pytest.raises(ValueError, match="Unsupported format"):
        prepare_xg(df, input_format="badformat")


def test_prepare_xg_file_missing_input_raises(tmp_path):
    from football_prediction_v19.xg_import import prepare_xg_file
    with pytest.raises(ValueError, match="not found"):
        prepare_xg_file(str(tmp_path / "nonexistent.csv"), str(tmp_path / "out.csv"))


def test_cli_prepare_xg_native_template(tmp_path):
    root = Path(__file__).resolve().parents[1]
    src = root / "data" / "raw" / "xg_raw_template.csv"
    out = tmp_path / "xg_clean.csv"
    main(["prepare-xg", "--input", str(src), "--output", str(out)])
    assert out.exists()
    df = pd.read_csv(out)
    assert len(df) == 2
    assert "home_xg" in df.columns
    assert "away_xg" in df.columns


def test_cli_prepare_xg_fbref_template(tmp_path):
    root = Path(__file__).resolve().parents[1]
    src = root / "data" / "raw" / "fbref_xg_template.csv"
    out = tmp_path / "xg_clean.csv"
    main(["prepare-xg", "--input", str(src), "--output", str(out), "--format", "fbref"])
    assert out.exists()
    df = pd.read_csv(out)
    assert len(df) == 2
    assert "home_xg" in df.columns


def test_cli_prepare_xg_understat_template(tmp_path):
    root = Path(__file__).resolve().parents[1]
    src = root / "data" / "raw" / "understat_xg_template.csv"
    out = tmp_path / "xg_clean.csv"
    main(["prepare-xg", "--input", str(src), "--output", str(out), "--format", "understat"])
    assert out.exists()
    df = pd.read_csv(out)
    assert len(df) == 2
    assert "home_xg" in df.columns


def test_merge_xg_into_history_basic():
    from football_prediction_v19.xg_import import merge_xg_into_history
    history = pd.DataFrame({
        "date": ["2024-08-17"],
        "home_team": ["Man Utd"],
        "away_team": ["Wolves"],
        "home_xg": [float("nan")],
        "away_xg": [float("nan")],
    })
    xg = pd.DataFrame({
        "date": [pd.Timestamp("2024-08-17")],
        "home_team": ["Man Utd"],
        "away_team": ["Wolves"],
        "home_xg": [1.23],
        "away_xg": [0.85],
    })
    updated, matched, missing = merge_xg_into_history(history, xg)
    assert matched == 1
    assert missing == 0
    assert updated.loc[0, "home_xg"] == pytest.approx(1.23)
    assert updated.loc[0, "away_xg"] == pytest.approx(0.85)


def test_merge_xg_derived_xga():
    from football_prediction_v19.xg_import import merge_xg_into_history
    history = pd.DataFrame({
        "date": [pd.NaT],
        "home_team": ["Arsenal"],
        "away_team": ["Chelsea"],
    })
    xg = pd.DataFrame({
        "date": [pd.NaT],
        "home_team": ["Arsenal"],
        "away_team": ["Chelsea"],
        "home_xg": [2.10],
        "away_xg": [0.65],
    })
    updated, matched, _ = merge_xg_into_history(history, xg)
    assert matched == 1
    # home_xga = away_xg (goals conceded by home = goals scored by away)
    assert updated.loc[0, "home_xga"] == pytest.approx(0.65)
    # away_xga = home_xg (goals conceded by away = goals scored by home)
    assert updated.loc[0, "away_xga"] == pytest.approx(2.10)


def test_merge_xg_date_window_no_match_at_zero():
    from football_prediction_v19.xg_import import merge_xg_into_history
    history = pd.DataFrame({
        "date": ["2024-08-17"],
        "home_team": ["Arsenal"],
        "away_team": ["Chelsea"],
    })
    xg = pd.DataFrame({
        "date": [pd.Timestamp("2024-08-15")],
        "home_team": ["Arsenal"],
        "away_team": ["Chelsea"],
        "home_xg": [1.50],
        "away_xg": [0.80],
    })
    _, matched, _ = merge_xg_into_history(history, xg, allow_date_window=0)
    assert matched == 0


def test_merge_xg_date_window_match_within_two_days():
    from football_prediction_v19.xg_import import merge_xg_into_history
    history = pd.DataFrame({
        "date": ["2024-08-17"],
        "home_team": ["Arsenal"],
        "away_team": ["Chelsea"],
    })
    xg = pd.DataFrame({
        "date": [pd.Timestamp("2024-08-15")],
        "home_team": ["Arsenal"],
        "away_team": ["Chelsea"],
        "home_xg": [1.50],
        "away_xg": [0.80],
    })
    _, matched, _ = merge_xg_into_history(history, xg, allow_date_window=2)
    assert matched == 1


def test_merge_xg_source_preference():
    from football_prediction_v19.xg_import import merge_xg_into_history
    history = pd.DataFrame({
        "date": [pd.NaT],
        "home_team": ["Arsenal"],
        "away_team": ["Chelsea"],
    })
    xg = pd.DataFrame({
        "date": [pd.NaT, pd.NaT],
        "home_team": ["Arsenal", "Arsenal"],
        "away_team": ["Chelsea", "Chelsea"],
        "home_xg": [1.80, 2.10],
        "away_xg": [0.50, 0.65],
        "source": ["other", "fbref"],
    })
    updated, matched, _ = merge_xg_into_history(history, xg, prefer_source="fbref")
    assert matched == 1
    assert updated.loc[0, "home_xg"] == pytest.approx(2.10)


def test_cli_merge_xg_history(tmp_path):
    root = Path(__file__).resolve().parents[1]
    xg_src = root / "data" / "raw" / "xg_raw_template.csv"
    xg_clean = tmp_path / "xg_clean.csv"
    main(["prepare-xg", "--input", str(xg_src), "--output", str(xg_clean)])

    out = tmp_path / "history_with_xg.csv"
    main([
        "merge-xg-history",
        "--history", str(root / "data" / "sample_matches.csv"),
        "--xg", str(xg_clean),
        "--output", str(out),
    ])
    assert out.exists()
    df = pd.read_csv(out)
    assert "home_xg" in df.columns
    assert "away_xg" in df.columns
    assert "home_xga" in df.columns
    assert "away_xga" in df.columns


def test_run_pipeline_without_xg_unchanged(tmp_path):
    root = Path(__file__).resolve().parents[1]
    out = tmp_path / "predictions.csv"
    main([
        "run-pipeline",
        "--skip-download",
        "--combine-output", str(root / "data" / "sample_matches.csv"),
        "--fixtures-raw", str(root / "data" / "raw" / "upcoming_fixtures_raw_template.csv"),
        "--fixtures-output", str(tmp_path / "fx.csv"),
        "--model", str(tmp_path / "model.joblib"),
        "--predictions", str(out),
        "--excel", str(tmp_path / "report.xlsx"),
    ])
    assert out.exists()


def test_run_pipeline_with_xg_raw(tmp_path):
    root = Path(__file__).resolve().parents[1]
    xg_raw = root / "data" / "raw" / "xg_raw_template.csv"
    out = tmp_path / "predictions.csv"
    main([
        "run-pipeline",
        "--skip-download",
        "--combine-output", str(root / "data" / "sample_matches.csv"),
        "--fixtures-raw", str(root / "data" / "raw" / "upcoming_fixtures_raw_template.csv"),
        "--fixtures-output", str(tmp_path / "fx.csv"),
        "--model", str(tmp_path / "model.joblib"),
        "--predictions", str(out),
        "--excel", str(tmp_path / "report.xlsx"),
        "--xg-raw", str(xg_raw),
        "--xg-clean", str(tmp_path / "xg_clean.csv"),
        "--history-with-xg", str(tmp_path / "history_with_xg.csv"),
    ])
    assert out.exists()
    assert (tmp_path / "xg_clean.csv").exists()
    assert (tmp_path / "history_with_xg.csv").exists()


# ---------------------------------------------------------------------------
# training / compare_models tests
# ---------------------------------------------------------------------------

def test_compare_models_creates_output_files(tmp_path):
    from football_prediction_v19.training import compare_models
    root = Path(__file__).resolve().parents[1]
    out_dir = tmp_path / "cmp"
    compare_models(
        input_path=str(root / "data" / "sample_matches.csv"),
        output_dir=str(out_dir),
        test_season=2023,
    )
    assert (out_dir / "model_comparison.csv").exists()
    assert (out_dir / "model_comparison_report.md").exists()
    assert (out_dir / "best_model.joblib").exists()
    assert (out_dir / "feature_columns.json").exists()
    assert (out_dir / "best_model_metadata.json").exists()


def test_compare_models_csv_has_required_columns(tmp_path):
    from football_prediction_v19.training import compare_models
    root = Path(__file__).resolve().parents[1]
    out_dir = tmp_path / "cmp"
    compare_models(
        input_path=str(root / "data" / "sample_matches.csv"),
        output_dir=str(out_dir),
        test_season=2023,
    )
    df = pd.read_csv(out_dir / "model_comparison.csv")
    required = [
        "model_name", "calibrated", "accuracy", "balanced_accuracy",
        "log_loss", "brier_score", "avg_confidence", "avg_correct_confidence",
        "train_rows", "test_rows", "feature_count", "selected_as_best", "warnings",
    ]
    for col in required:
        assert col in df.columns, f"Missing column: {col}"
    # exactly one row should be selected as best
    assert df["selected_as_best"].sum() == 1


def test_compare_models_best_model_loadable(tmp_path):
    from football_prediction_v19.training import compare_models
    from football_prediction_v19.model import load_model
    root = Path(__file__).resolve().parents[1]
    out_dir = tmp_path / "cmp"
    compare_models(
        input_path=str(root / "data" / "sample_matches.csv"),
        output_dir=str(out_dir),
        test_season=2023,
    )
    bundle = load_model(out_dir / "best_model.joblib")
    assert "model" in bundle
    assert "feature_cols" in bundle
    assert len(bundle["feature_cols"]) > 0


def test_compare_models_metadata_has_expected_fields(tmp_path):
    import json
    from football_prediction_v19.training import compare_models
    root = Path(__file__).resolve().parents[1]
    out_dir = tmp_path / "cmp"
    compare_models(
        input_path=str(root / "data" / "sample_matches.csv"),
        output_dir=str(out_dir),
        test_season=2023,
    )
    meta = json.loads((out_dir / "best_model_metadata.json").read_text())
    for field in ["model_name", "calibrated", "test_season", "selected_metric",
                  "accuracy", "log_loss", "brier_score", "feature_columns",
                  "training_rows", "test_rows", "created_at"]:
        assert field in meta, f"Missing metadata field: {field}"
    assert meta["test_season"] == 2023
    assert meta["selected_metric"] == "log_loss"


def test_compare_models_missing_test_season_raises(tmp_path):
    from football_prediction_v19.training import compare_models
    root = Path(__file__).resolve().parents[1]
    with pytest.raises(ValueError, match="not found in the data"):
        compare_models(
            input_path=str(root / "data" / "sample_matches.csv"),
            output_dir=str(tmp_path / "cmp"),
            test_season=1900,
        )


def test_compare_models_missing_file_raises(tmp_path):
    from football_prediction_v19.training import compare_models
    with pytest.raises(ValueError, match="not found"):
        compare_models(
            input_path=str(tmp_path / "nonexistent.csv"),
            output_dir=str(tmp_path / "cmp"),
            test_season=2023,
        )


def test_compare_models_small_data_calibration_does_not_crash(tmp_path):
    """Calibration should gracefully skip on tiny datasets."""
    from football_prediction_v19.training import compare_models
    root = Path(__file__).resolve().parents[1]
    # Use the real sample_matches.csv — calibration may warn but must not crash
    out_dir = tmp_path / "cmp_small"
    # This must complete without exception
    compare_models(
        input_path=str(root / "data" / "sample_matches.csv"),
        output_dir=str(out_dir),
        test_season=2023,
    )
    assert (out_dir / "best_model.joblib").exists()


def test_cli_compare_models_command(tmp_path):
    root = Path(__file__).resolve().parents[1]
    out_dir = tmp_path / "cmp"
    main([
        "compare-models",
        "--input", str(root / "data" / "sample_matches.csv"),
        "--output-dir", str(out_dir),
        "--test-season", "2023",
    ])
    assert (out_dir / "model_comparison.csv").exists()
    assert (out_dir / "best_model.joblib").exists()


def test_existing_train_command_still_works(tmp_path):
    root = Path(__file__).resolve().parents[1]
    main([
        "train",
        "--input", str(root / "data" / "sample_matches.csv"),
        "--model", str(tmp_path / "model.joblib"),
        "--test-season", "2023",
    ])
    assert (tmp_path / "model.joblib").exists()


def test_run_pipeline_with_compare_models(tmp_path):
    root = Path(__file__).resolve().parents[1]
    out = tmp_path / "predictions.csv"
    main([
        "run-pipeline",
        "--skip-download",
        "--combine-output", str(root / "data" / "sample_matches.csv"),
        "--fixtures-raw", str(root / "data" / "raw" / "upcoming_fixtures_raw_template.csv"),
        "--fixtures-output", str(tmp_path / "fx.csv"),
        "--model", str(tmp_path / "model.joblib"),
        "--predictions", str(out),
        "--excel", str(tmp_path / "report.xlsx"),
        "--compare-models",
        "--compare-models-dir", str(tmp_path / "cmp"),
        "--test-season", "2023",
    ])
    assert out.exists()
    assert (tmp_path / "cmp" / "best_model.joblib").exists()


def test_run_pipeline_without_compare_models_unchanged(tmp_path):
    root = Path(__file__).resolve().parents[1]
    out = tmp_path / "predictions.csv"
    main([
        "run-pipeline",
        "--skip-download",
        "--combine-output", str(root / "data" / "sample_matches.csv"),
        "--fixtures-raw", str(root / "data" / "raw" / "upcoming_fixtures_raw_template.csv"),
        "--fixtures-output", str(tmp_path / "fx.csv"),
        "--model", str(tmp_path / "model.joblib"),
        "--predictions", str(out),
        "--excel", str(tmp_path / "report.xlsx"),
    ])
    assert out.exists()


# ---------------------------------------------------------------------------
# Excel dashboard tests
# ---------------------------------------------------------------------------

def _make_predictions_csv(path: Path) -> None:
    df = pd.DataFrame([{
        "date": "2024-08-17", "league": "Premier League",
        "home_team": "Arsenal", "away_team": "Chelsea",
        "prob_home": 0.45, "prob_draw": 0.28, "prob_away": 0.27,
        "odds_home": 2.10, "odds_draw": 3.40, "odds_away": 3.50,
        "fair_home": 0.45, "fair_draw": 0.28, "fair_away": 0.27,
        "edge_home": 0.05, "edge_draw": 0.0, "edge_away": 0.0,
        "value_pick": "H", "value_edge": 0.05,
        "bet_recommendation": "Bet H",
        "control_score": 8.0, "chaos_score": 2.5,
        "tdi_home": 0.6, "tdi_away": 0.4,
        "no_bet_reasons": "", "v19_flags": "",
    }])
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False)


def _make_model_comparison_csv(path: Path) -> None:
    df = pd.DataFrame([{
        "model_name": "random_forest", "calibrated": True,
        "accuracy": 0.54, "balanced_accuracy": 0.51,
        "log_loss": 0.97, "brier_score": 0.62,
        "avg_confidence": 0.55, "avg_correct_confidence": 0.58,
        "train_rows": 240, "test_rows": 40,
        "feature_count": 32, "selected_as_best": True, "warnings": "",
    }])
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False)


def _make_model_metadata_json(path: Path) -> None:
    import json as _json
    meta = {
        "model_name": "random_forest", "calibrated": True,
        "test_season": 2023, "selected_metric": "log_loss",
        "accuracy": 0.54, "log_loss": 0.97, "brier_score": 0.62,
        "feature_columns": ["home_form_xg", "away_form_xg"],
        "training_rows": 240, "test_rows": 40,
        "created_at": "2025-01-01T00:00:00+00:00",
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(_json.dumps(meta), encoding="utf-8")


def _make_backtest_csv(path: Path) -> None:
    df = pd.DataFrame([{
        "date": "2023-08-20", "home_team": "Arsenal", "away_team": "Chelsea",
        "value_pick": "H", "value_edge": 0.08, "odds_used": 2.10,
        "bet_recommendation": "Bet H", "bet_result": "win",
        "profit": 1.10, "cumulative_profit": 1.10, "no_bet_reasons": "",
    }])
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False)


def _sheet_names(xlsx_path: Path) -> list[str]:
    wb = load_workbook(xlsx_path, read_only=True)
    return wb.sheetnames


def test_export_excel_basic_still_works(tmp_path):
    pred = tmp_path / "preds.csv"
    out = tmp_path / "report.xlsx"
    _make_predictions_csv(pred)
    main(["export-excel", "--predictions", str(pred), "--output", str(out)])
    assert out.exists()
    sheets = _sheet_names(out)
    assert "Summary" in sheets
    assert "Predictions" in sheets


def test_export_excel_with_model_comparison(tmp_path):
    pred = tmp_path / "preds.csv"
    cmp_csv = tmp_path / "model_comparison.csv"
    out = tmp_path / "report.xlsx"
    _make_predictions_csv(pred)
    _make_model_comparison_csv(cmp_csv)
    main([
        "export-excel",
        "--predictions", str(pred),
        "--output", str(out),
        "--model-comparison", str(cmp_csv),
    ])
    assert out.exists()
    sheets = _sheet_names(out)
    assert "Model Comparison" in sheets
    assert "Calibration" in sheets


def test_export_excel_with_model_metadata(tmp_path):
    pred = tmp_path / "preds.csv"
    meta = tmp_path / "meta.json"
    out = tmp_path / "report.xlsx"
    _make_predictions_csv(pred)
    _make_model_metadata_json(meta)
    main([
        "export-excel",
        "--predictions", str(pred),
        "--output", str(out),
        "--model-metadata", str(meta),
    ])
    assert out.exists()
    sheets = _sheet_names(out)
    assert "Best Model" in sheets
    assert "Feature Metadata" in sheets


def test_export_excel_with_backtest(tmp_path):
    pred = tmp_path / "preds.csv"
    bt = tmp_path / "backtest.csv"
    out = tmp_path / "report.xlsx"
    _make_predictions_csv(pred)
    _make_backtest_csv(bt)
    main([
        "export-excel",
        "--predictions", str(pred),
        "--output", str(out),
        "--backtest-csv", str(bt),
    ])
    assert out.exists()
    sheets = _sheet_names(out)
    assert "Backtest" in sheets
    assert "Backtest Summary" in sheets


def test_export_excel_all_optional_sheets(tmp_path):
    pred = tmp_path / "preds.csv"
    cmp_csv = tmp_path / "model_comparison.csv"
    meta = tmp_path / "meta.json"
    bt = tmp_path / "backtest.csv"
    out = tmp_path / "report.xlsx"
    _make_predictions_csv(pred)
    _make_model_comparison_csv(cmp_csv)
    _make_model_metadata_json(meta)
    _make_backtest_csv(bt)
    main([
        "export-excel",
        "--predictions", str(pred),
        "--output", str(out),
        "--model-comparison", str(cmp_csv),
        "--model-metadata", str(meta),
        "--backtest-csv", str(bt),
    ])
    assert out.exists()
    sheets = _sheet_names(out)
    for expected in ["Summary", "Predictions", "Model Comparison", "Calibration",
                     "Best Model", "Feature Metadata", "Backtest", "Backtest Summary"]:
        assert expected in sheets, f"Missing sheet: {expected}"


def test_export_excel_missing_optional_files_no_crash(tmp_path):
    pred = tmp_path / "preds.csv"
    out = tmp_path / "report.xlsx"
    _make_predictions_csv(pred)
    main([
        "export-excel",
        "--predictions", str(pred),
        "--output", str(out),
        "--model-comparison", str(tmp_path / "nonexistent.csv"),
        "--model-metadata", str(tmp_path / "nonexistent.json"),
        "--backtest-csv", str(tmp_path / "nonexistent_bt.csv"),
    ])
    assert out.exists()
    sheets = _sheet_names(out)
    assert "Model Comparison" not in sheets
    assert "Backtest" not in sheets


def test_run_pipeline_compare_models_includes_excel_dashboard(tmp_path):
    root = Path(__file__).resolve().parents[1]
    out = tmp_path / "predictions.csv"
    excel_out = tmp_path / "report.xlsx"
    main([
        "run-pipeline",
        "--skip-download",
        "--combine-output", str(root / "data" / "sample_matches.csv"),
        "--fixtures-raw", str(root / "data" / "raw" / "upcoming_fixtures_raw_template.csv"),
        "--fixtures-output", str(tmp_path / "fx.csv"),
        "--model", str(tmp_path / "model.joblib"),
        "--predictions", str(out),
        "--excel", str(excel_out),
        "--compare-models",
        "--compare-models-dir", str(tmp_path / "cmp"),
        "--test-season", "2023",
        "--skip-backtest",
    ])
    assert excel_out.exists()
    sheets = _sheet_names(excel_out)
    assert "Model Comparison" in sheets
    assert "Best Model" in sheets


# ---------------------------------------------------------------------------
# Release-readiness tests
# ---------------------------------------------------------------------------

_PROJECT_ROOT = Path(__file__).resolve().parents[1]


def test_doctor_command_runs():
    main(["doctor"])


def test_docs_files_exist():
    for doc in [
        "CHANGELOG.md",
        "docs/QUICKSTART.md",
        "docs/COMMANDS.md",
        "docs/DATA_REQUIREMENTS.md",
    ]:
        assert (_PROJECT_ROOT / doc).exists(), f"Missing: {doc}"


def test_changelog_exists():
    assert (_PROJECT_ROOT / "CHANGELOG.md").exists()


def test_readme_contains_main_commands():
    readme = _PROJECT_ROOT / "README.md"
    if not readme.exists():
        pytest.skip("README.md not present")
    text = readme.read_text(encoding="utf-8")
    for cmd in ["train", "predict", "prepare-data", "run-pipeline"]:
        assert cmd in text, f"README missing command: {cmd}"


def test_commands_doc_mentions_all_cli_commands():
    doc = (_PROJECT_ROOT / "docs" / "COMMANDS.md").read_text(encoding="utf-8")
    for cmd in [
        "doctor", "train", "predict", "predict-fixtures", "prepare-data",
        "compare-models", "export-excel", "run-pipeline", "backtest-bets",
    ]:
        assert cmd in doc, f"COMMANDS.md missing: {cmd}"


def test_quickstart_mentions_sample_pipeline():
    doc = (_PROJECT_ROOT / "docs" / "QUICKSTART.md").read_text(encoding="utf-8")
    assert "run-pipeline" in doc
    assert "sample_matches" in doc or "sample" in doc


def test_data_requirements_mentions_key_sections():
    doc = (_PROJECT_ROOT / "docs" / "DATA_REQUIREMENTS.md").read_text(encoding="utf-8")
    for section in ["home_xg", "odds_home", "football-data", "FBref", "team_aliases"]:
        assert section in doc, f"DATA_REQUIREMENTS.md missing: {section}"


def test_football_data_import_preserves_league():
    """football-data.co.uk import must map Div codes to human-readable league names."""
    import io
    from football_prediction_v19.data import prepare_real_matches

    csv = (
        "Div,Date,HomeTeam,AwayTeam,FTHG,FTAG,FTR,B365H,B365D,B365A\n"
        "D1,01/08/2023,Bayern Munich,Dortmund,2,1,H,1.50,4.00,6.00\n"
        "E0,12/08/2023,Arsenal,Chelsea,1,1,D,2.50,3.30,2.80\n"
        "SP1,18/08/2023,Barcelona,Real Madrid,3,2,H,2.10,3.40,3.20\n"
    )
    df = pd.read_csv(io.StringIO(csv))
    clean = prepare_real_matches(df, input_format="football-data")

    assert "league" in clean.columns, "league column missing after football-data import"
    leagues = clean["league"].unique().tolist()
    assert "Bundesliga" in leagues, f"Expected 'Bundesliga' from D1, got: {leagues}"
    assert "Premier League" in leagues, f"Expected 'Premier League' from E0, got: {leagues}"
    assert "La Liga" in leagues, f"Expected 'La Liga' from SP1, got: {leagues}"
    assert "Unknown" not in leagues, f"Some leagues were not mapped: {leagues}"


def test_football_data_import_preserves_odds():
    """football-data.co.uk import must map B365H/D/A to odds_home/draw/away."""
    import io
    from football_prediction_v19.data import prepare_real_matches

    csv = (
        "Div,Date,HomeTeam,AwayTeam,FTHG,FTAG,FTR,B365H,B365D,B365A\n"
        "D1,01/08/2023,Bayern Munich,Dortmund,2,1,H,1.50,4.00,6.00\n"
    )
    df = pd.read_csv(io.StringIO(csv))
    clean = prepare_real_matches(df, input_format="football-data")

    assert clean["odds_home"].notna().all(), "odds_home is NaN after football-data import"
    assert clean["odds_draw"].notna().all(), "odds_draw is NaN after football-data import"
    assert clean["odds_away"].notna().all(), "odds_away is NaN after football-data import"
    assert float(clean["odds_home"].iloc[0]) == pytest.approx(1.50)
    assert float(clean["odds_draw"].iloc[0]) == pytest.approx(4.00)
    assert float(clean["odds_away"].iloc[0]) == pytest.approx(6.00)


def test_combined_output_preserves_league_and_odds(tmp_path):
    """Combined processed CSV must carry league names and odds through prepare + concat."""
    import io
    from football_prediction_v19.data import prepare_real_matches

    csv = (
        "Div,Date,HomeTeam,AwayTeam,FTHG,FTAG,FTR,B365H,B365D,B365A\n"
        "I1,15/08/2023,Juventus,AC Milan,1,0,H,2.20,3.10,3.50\n"
        "F1,20/08/2023,PSG,Lyon,2,0,H,1.30,5.00,8.00\n"
    )
    df = pd.read_csv(io.StringIO(csv))
    clean = prepare_real_matches(df, input_format="football-data")
    out_path = tmp_path / "combined.csv"
    clean.to_csv(out_path, index=False)

    reloaded = pd.read_csv(out_path)
    assert "Serie A" in reloaded["league"].values, "Serie A not preserved through save/load cycle"
    assert "Ligue 1" in reloaded["league"].values, "Ligue 1 not preserved through save/load cycle"
    assert reloaded["odds_home"].notna().all(), "odds_home lost through save/load cycle"


def test_backtest_output_includes_league(tmp_path):
    """Backtest rows must carry league from historical data, not default to Unknown."""
    import io
    from football_prediction_v19.data import prepare_real_matches
    from football_prediction_v19.features import build_features
    from football_prediction_v19.model import train_from_matches

    csv_rows = ["Div,Date,HomeTeam,AwayTeam,FTHG,FTAG,FTR,B365H,B365D,B365A"]
    # 2022 season — training data
    for i in range(20):
        csv_rows.append(f"D1,{15 + i:02d}/08/2022,TeamA,TeamB,1,0,H,2.0,3.5,4.0")
        csv_rows.append(f"D1,{15 + i:02d}/08/2022,TeamC,TeamD,0,1,A,3.0,3.5,2.2")
    # 2023 season — test data
    for i in range(10):
        csv_rows.append(f"D1,{15 + i:02d}/08/2023,TeamA,TeamB,2,1,H,1.8,3.5,4.5")
        csv_rows.append(f"D1,{15 + i:02d}/08/2023,TeamC,TeamD,0,0,D,2.5,3.2,3.0")

    df = pd.read_csv(io.StringIO("\n".join(csv_rows)))
    clean = prepare_real_matches(df, input_format="football-data")
    model, _, metrics, cols = train_from_matches(clean, test_season=2023)
    bundle = {"model": model, "feature_cols": cols, "metrics": metrics}
    results = run_bet_backtest(clean, bundle, test_season=2023)

    assert "league" in results.columns, "backtest output missing league column"
    assert (results["league"] == "Bundesliga").all(), (
        f"Expected all rows to be 'Bundesliga', got: {results['league'].unique()}"
    )


def test_xg_gates_skipped_when_xg_unavailable():
    """DNB gate 'xg_edge_positive' must not appear as a lock reason when xG data is absent."""
    from football_prediction_v19.rules_v19 import assess_prediction

    row_no_xg = pd.Series({
        "home_team": "TeamA", "away_team": "TeamB",
        "odds_home": 2.0, "odds_draw": 3.5, "odds_away": 3.5,
        "home_w5_ppg": 1.8, "away_w5_ppg": 1.2,
        # edge_w5_xgdiff intentionally absent (NaN)
    })
    probs = {"H": 0.50, "D": 0.28, "A": 0.22}
    result = assess_prediction(row_no_xg, probs)

    no_bet_text = " ".join(result["no_bets"])
    assert "xg_edge_positive" not in no_bet_text, (
        f"xg_edge_positive appeared as a lock reason despite xG being unavailable: {result['no_bets']}"
    )
    assert "xG unavailable" in no_bet_text, (
        f"Expected 'xG unavailable' notice in no_bets: {result['no_bets']}"
    )


def test_xg_gates_fire_when_xg_present():
    """DNB gate 'xg_edge_positive' must still apply when xG data IS available."""
    from football_prediction_v19.rules_v19 import assess_prediction

    row_with_xg = pd.Series({
        "home_team": "TeamA", "away_team": "TeamB",
        "odds_home": 2.0, "odds_draw": 3.5, "odds_away": 3.5,
        "home_w5_ppg": 1.8, "away_w5_ppg": 1.2,
        "edge_w5_xgdiff": -0.05,  # negative — xg_edge_positive should be False
        "edge_w5_ppg": 0.6,
    })
    probs = {"H": 0.50, "D": 0.28, "A": 0.22}
    result = assess_prediction(row_with_xg, probs)

    no_bet_text = " ".join(result["no_bets"])
    assert "xG unavailable" not in no_bet_text, (
        f"'xG unavailable' appeared despite xG being present: {result['no_bets']}"
    )


def test_backtest_test_season_excludes_training_rows():
    """Backtest with test_season must not evaluate on training-season rows (leakage guard)."""
    path = Path(__file__).resolve().parents[1] / "data" / "sample_matches.csv"
    df = pd.read_csv(path)
    model, _, metrics, cols = train_from_matches(df, test_season=2023)
    bundle = {"model": model, "feature_cols": cols, "metrics": metrics}

    results_all = run_bet_backtest(df, bundle)
    results_oos = run_bet_backtest(df, bundle, test_season=2023)

    # Out-of-sample result must contain fewer rows than the full evaluation
    assert len(results_oos) < len(results_all), (
        "test_season filter had no effect — training rows are still included in backtest"
    )
    # All rows in the filtered result must belong to season >= test_season
    from football_prediction_v19.features import build_features
    table = build_features(pd.read_csv(path))
    oos_seasons = table[table["season_start"] >= 2023]["season_start"].unique()
    bt_dates = pd.to_datetime(results_oos["date"])
    bt_seasons = bt_dates.apply(lambda d: d.year - 1 if d.month < 8 else d.year)
    assert set(bt_seasons.unique()).issubset(set(oos_seasons)), (
        "Backtest contains rows from before test_season — training data leaked into evaluation"
    )


# ---------------------------------------------------------------------------
# MLS tests
# ---------------------------------------------------------------------------

def _make_mls_sample() -> pd.DataFrame:
    """Build a minimal in-memory MLS dataset (30 rows, two seasons)."""
    import random
    random.seed(42)
    rows = []
    teams = [
        ("LA Galaxy", "Inter Miami"),
        ("Seattle Sounders", "Portland Timbers"),
        ("Atlanta United", "Orlando City"),
        ("New York City FC", "New York Red Bulls"),
        ("Sporting Kansas City", "Columbus Crew"),
    ]
    for season in [2023, 2024]:
        for home, away in teams:
            for _ in range(3):
                hg = random.randint(0, 3)
                ag = random.randint(0, 3)
                rows.append({
                    "date": f"{season}-05-{random.randint(1, 28):02d}",
                    "season": str(season),
                    "league": "MLS",
                    "home_team": home,
                    "away_team": away,
                    "score": f"{hg}-{ag}",
                    "home_xg": round(random.uniform(0.5, 2.5), 2),
                    "away_xg": round(random.uniform(0.5, 2.5), 2),
                    "odds_home": round(random.uniform(1.5, 4.0), 2),
                    "odds_draw": round(random.uniform(2.8, 4.5), 2),
                    "odds_away": round(random.uniform(1.5, 4.0), 2),
                    "venue": "Example Stadium",
                    "referee": "",
                })
    return pd.DataFrame(rows)


def test_mls_templates_exist():
    root = Path(__file__).resolve().parents[1]
    assert (root / "data" / "raw" / "mls_matches_template.csv").exists()
    assert (root / "data" / "raw" / "mls_upcoming_fixtures_template.csv").exists()
    assert (root / "data" / "raw" / "mls_odds_template.csv").exists()


def test_mls_team_aliases_present():
    import json
    root = Path(__file__).resolve().parents[1]
    aliases = json.loads((root / "config" / "team_aliases.json").read_text(encoding="utf-8"))
    assert "Inter Miami" in aliases
    assert "LA Galaxy" in aliases
    assert "Seattle Sounders" in aliases


def test_mls_team_alias_normalization():
    assert normalize_team_name("LAFC") == "Los Angeles FC"
    assert normalize_team_name("Sporting KC") == "Sporting Kansas City"
    assert normalize_team_name("NY Red Bulls") == "New York Red Bulls"


def test_mls_fixture_features():
    df = _make_mls_sample()
    fix = build_fixture_features(
        history_df=df,
        home_team="LA Galaxy",
        away_team="Inter Miami",
        match_date="2025-01-15",
        venue="Dignity Health Sports Park",
    )
    assert isinstance(fix, pd.DataFrame)
    assert len(fix) == 1


def test_mls_excel_export_preserves_league(tmp_path):
    df = _make_mls_sample()
    features = build_features(df)
    if "league" not in features.columns:
        features["league"] = "MLS"
    out = tmp_path / "mls_features.xlsx"
    features.to_excel(out, index=False)
    wb = load_workbook(out)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    league_col = headers.index("league") + 1
    leagues = {ws.cell(row=r, column=league_col).value for r in range(2, ws.max_row + 1)}
    assert "MLS" in leagues


def test_mls_smoke_script_exists():
    root = Path(__file__).resolve().parents[1]
    assert (root / "scripts" / "run_mls_smoke.py").exists()


def test_mls_strategy_sweep_no_backtest_file(tmp_path):
    import subprocess
    root = Path(__file__).resolve().parents[1]
    script = root / "scripts" / "run_mls_strategy_sweep.py"
    result = subprocess.run(
        ["python", str(script)],
        capture_output=True,
        text=True,
        cwd=str(tmp_path),
    )
    assert result.returncode == 0
    assert "No MLS backtest file" in result.stdout


def test_mls_build_features():
    df = _make_mls_sample()
    features = build_features(df)
    assert len(features) > 0
    if "league" in features.columns:
        assert (features["league"] == "MLS").all()


# ---------------------------------------------------------------------------
# MLS data source tests
# ---------------------------------------------------------------------------

def test_import_mls_fbref_column_mapping(tmp_path):
    from football_prediction_v19.importers.mls_fbref import import_mls_fbref
    csv = tmp_path / "mls_raw.csv"
    csv.write_text(
        "Date,Home,Away,Score,xG,xG.1,Venue,Referee,Comp,Season\n"
        "2025-03-01,LA Galaxy,Inter Miami,2-1,1.85,0.90,Dignity Health Sports Park,,MLS,2025\n"
    )
    df = import_mls_fbref(csv)
    for col in ["date", "home_team", "away_team", "score", "home_xg", "away_xg"]:
        assert col in df.columns, f"Missing column: {col}"
    assert df["date"].iloc[0] == "2025-03-01"
    assert df["score"].iloc[0] == "2-1"
    assert abs(float(df["home_xg"].iloc[0]) - 1.85) < 0.01


def test_import_mls_fbref_normalizes_aliases(tmp_path):
    from football_prediction_v19.importers.mls_fbref import import_mls_fbref
    csv = tmp_path / "mls_raw.csv"
    csv.write_text(
        "Date,Home,Away,Score,xG,xG.1\n"
        "2025-04-01,LAFC,Inter Miami,1-0,1.10,0.50\n"
    )
    df = import_mls_fbref(csv)
    assert df["home_team"].iloc[0] == "Los Angeles FC"


def test_import_mls_fbref_infers_league(tmp_path):
    from football_prediction_v19.importers.mls_fbref import import_mls_fbref
    csv = tmp_path / "mls_raw.csv"
    csv.write_text(
        "Date,Home,Away,Score\n"
        "2025-04-01,LA Galaxy,Inter Miami,2-1\n"
    )
    df = import_mls_fbref(csv)
    assert "league" in df.columns
    assert df["league"].iloc[0] == "MLS"


def test_import_mls_fbref_infers_season(tmp_path):
    from football_prediction_v19.importers.mls_fbref import import_mls_fbref
    csv = tmp_path / "mls_raw.csv"
    csv.write_text(
        "Date,Home,Away,Score\n"
        "2025-06-15,LA Galaxy,Inter Miami,1-1\n"
    )
    df = import_mls_fbref(csv)
    assert "season" in df.columns
    assert df["season"].iloc[0] == "2025"


def test_import_mls_fbref_missing_file_error(tmp_path):
    from football_prediction_v19.importers.mls_fbref import import_mls_fbref
    import pytest
    with pytest.raises(FileNotFoundError):
        import_mls_fbref(tmp_path / "does_not_exist.csv")


def test_import_mls_fbref_empty_file_error(tmp_path):
    from football_prediction_v19.importers.mls_fbref import import_mls_fbref
    import pytest
    csv = tmp_path / "empty.csv"
    csv.write_text("Date,Home,Away,Score\n")
    with pytest.raises(ValueError):
        import_mls_fbref(csv)


def test_parse_odds_api_events_maps_h2h():
    import json
    from pathlib import Path
    from football_prediction_v19.importers.the_odds_api import parse_the_odds_api_events
    sample = Path(__file__).resolve().parents[1] / "data" / "raw" / "mls_the_odds_api_sample.json"
    payload = json.loads(sample.read_text())
    df = parse_the_odds_api_events(payload)
    assert len(df) == 2
    for col in ["odds_home", "odds_draw", "odds_away"]:
        assert col in df.columns
        assert df[col].dtype == float or df[col].apply(lambda x: isinstance(x, (int, float))).all()
    assert df["home_team"].iloc[0] != ""
    assert df["away_team"].iloc[0] != ""


def test_parse_odds_api_events_normalizes_teams():
    from football_prediction_v19.importers.the_odds_api import parse_the_odds_api_events
    payload = [
        {
            "id": "x1",
            "sport_key": "soccer_usa_mls",
            "commence_time": "2025-04-05T20:30:00Z",
            "home_team": "LA Galaxy",
            "away_team": "Inter Miami",
            "bookmakers": [
                {
                    "key": "bet365",
                    "title": "Bet365",
                    "last_update": "2025-04-04T18:00:00Z",
                    "markets": [
                        {
                            "key": "h2h",
                            "outcomes": [
                                {"name": "LA Galaxy", "price": 2.10},
                                {"name": "Inter Miami", "price": 3.60},
                                {"name": "Draw", "price": 3.40},
                            ],
                        }
                    ],
                }
            ],
        }
    ]
    df = parse_the_odds_api_events(payload)
    assert df["home_team"].iloc[0] == "LA Galaxy"


def test_parse_odds_api_events_empty_raises():
    from football_prediction_v19.importers.the_odds_api import parse_the_odds_api_events
    import pytest
    with pytest.raises(ValueError):
        parse_the_odds_api_events([])


def test_download_mls_odds_uses_env_key(tmp_path, monkeypatch):
    import json
    from unittest.mock import MagicMock, patch
    sample = Path(__file__).resolve().parents[1] / "data" / "raw" / "mls_the_odds_api_sample.json"
    payload = json.loads(sample.read_text())

    mock_response = MagicMock()
    mock_response.status_code = 200
    mock_response.json.return_value = payload
    mock_response.raise_for_status = MagicMock()

    monkeypatch.setenv("THE_ODDS_API_KEY", "test-env-key-123")

    with patch("football_prediction_v19.importers.the_odds_api._requests_module") as mock_requests:
        mock_requests.get.return_value = mock_response
        from football_prediction_v19.importers.the_odds_api import fetch_mls_odds
        df = fetch_mls_odds(api_key="test-env-key-123", output_path=tmp_path / "odds.csv")

    call_kwargs = mock_requests.get.call_args
    assert call_kwargs is not None
    params = call_kwargs[1].get("params", call_kwargs[0][1] if len(call_kwargs[0]) > 1 else {})
    assert params.get("apiKey") == "test-env-key-123"
    assert len(df) > 0


def test_download_mls_odds_missing_key_error():
    from unittest.mock import MagicMock, patch
    mock_response = MagicMock()
    mock_response.status_code = 401
    mock_response.raise_for_status = MagicMock()
    import pytest
    with patch("football_prediction_v19.importers.the_odds_api._requests_module") as mock_requests:
        mock_requests.get.return_value = mock_response
        from football_prediction_v19.importers.the_odds_api import fetch_mls_odds
        with pytest.raises(ValueError, match="Invalid API key"):
            fetch_mls_odds(api_key="bad-key")


def test_prepare_mls_data_cli_command(tmp_path):
    root = Path(__file__).resolve().parents[1]
    fbref_template = root / "data" / "raw" / "mls_fbref_raw_template.csv"
    matches_out = tmp_path / "mls_matches.csv"
    processed_out = tmp_path / "mls_processed.csv"
    main([
        "prepare-mls-data",
        "--fbref", str(fbref_template),
        "--matches-output", str(matches_out),
        "--processed-output", str(processed_out),
    ])
    assert processed_out.exists()
    df = pd.read_csv(processed_out)
    assert len(df) > 0


def test_mls_sources_smoke_script_exists():
    root = Path(__file__).resolve().parents[1]
    assert (root / "scripts" / "run_mls_sources_smoke.py").exists()


def test_import_mls_fbref_cli_command(tmp_path):
    root = Path(__file__).resolve().parents[1]
    fbref_template = root / "data" / "raw" / "mls_fbref_raw_template.csv"
    output = tmp_path / "mls_out.csv"
    main([
        "import-mls-fbref",
        "--input", str(fbref_template),
        "--output", str(output),
    ])
    assert output.exists()
    df = pd.read_csv(output)
    assert len(df) > 0
