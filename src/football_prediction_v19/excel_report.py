from __future__ import annotations

import json
from collections import Counter
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

PREDICTION_COLUMNS = [
    "date",
    "league",
    "home_team",
    "away_team",
    "prob_home",
    "prob_draw",
    "prob_away",
    "odds_home",
    "odds_draw",
    "odds_away",
    "fair_home",
    "fair_draw",
    "fair_away",
    "edge_home",
    "edge_draw",
    "edge_away",
    "value_pick",
    "value_edge",
    "bet_recommendation",
    "control_score",
    "chaos_score",
    "tdi_home",
    "tdi_away",
    "no_bet_reasons",
    "v19_flags",
]

PERCENT_COLUMNS = {
    "prob_home",
    "prob_draw",
    "prob_away",
    "fair_home",
    "fair_draw",
    "fair_away",
    "edge_home",
    "edge_draw",
    "edge_away",
    "value_edge",
}

ODDS_COLUMNS = {"odds_home", "odds_draw", "odds_away"}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_BEST_FILL = PatternFill("solid", fgColor="E2EFDA")   # light green for best model row
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)


def _ensure_columns(df: pd.DataFrame, columns: Iterable[str]) -> pd.DataFrame:
    out = df.copy()
    for col in columns:
        if col not in out.columns:
            out[col] = pd.NA
    return out[list(columns)]


def _split_reasons(series: pd.Series) -> list[str]:
    reasons: list[str] = []
    for text in series.fillna(""):
        for reason in str(text).split(" | "):
            reason = reason.strip()
            if reason:
                reasons.append(reason)
    return reasons


def _write_rows(ws, rows: list[list[object]]) -> None:
    for row in rows:
        ws.append(row)


def _format_sheet(
    ws,
    percent_columns: set[str] | None = None,
    odds_columns: set[str] | None = None,
    highlight_col: str | None = None,
    highlight_fill: PatternFill | None = None,
) -> None:
    percent_columns = percent_columns or set()
    odds_columns = odds_columns or set()
    if ws.max_row == 0:
        return
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    headers = [cell.value for cell in ws[1]]
    highlight_idx: int | None = None
    if highlight_col and highlight_col in headers:
        highlight_idx = headers.index(highlight_col)
    for idx, header in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        width = min(max(len(str(header or "")) + 2, 12), 38)
        for cell in ws[letter][1:]:
            width = min(max(width, len(str(cell.value or "")) + 2), 48)
        ws.column_dimensions[letter].width = width
        if header in percent_columns:
            for cell in ws[letter][1:]:
                cell.number_format = "0.0%"
        if header in odds_columns:
            for cell in ws[letter][1:]:
                cell.number_format = "0.00"

    # Highlight rows where highlight_col is truthy
    if highlight_idx is not None and highlight_fill is not None:
        for row_cells in ws.iter_rows(min_row=2):
            val = row_cells[highlight_idx].value
            if val and str(val).lower() not in ("false", "0", "no", ""):
                for cell in row_cells:
                    cell.fill = highlight_fill


def _add_dataframe_sheet(
    wb: Workbook,
    name: str,
    df: pd.DataFrame,
    percent_cols: set[str] | None = None,
    odds_cols: set[str] | None = None,
    highlight_col: str | None = None,
    highlight_fill: PatternFill | None = None,
) -> None:
    ws = wb.create_sheet(name)
    rows = [list(df.columns)] + df.where(pd.notna(df), None).values.tolist()
    _write_rows(ws, rows)
    _format_sheet(
        ws,
        percent_columns=percent_cols or PERCENT_COLUMNS,
        odds_columns=odds_cols or ODDS_COLUMNS,
        highlight_col=highlight_col,
        highlight_fill=highlight_fill,
    )


def _summary_rows(df: pd.DataFrame) -> list[list[object]]:
    recommended = df[df["bet_recommendation"].fillna("").str.lower() != "no bet"]
    no_bets = df[df["bet_recommendation"].fillna("").str.lower() == "no bet"]
    avg_edge = recommended["value_edge"].mean() if not recommended.empty else 0
    rows: list[list[object]] = [
        ["Metric", "Value"],
        ["Total fixtures", len(df)],
        ["Recommended bets", len(recommended)],
        ["No-bets", len(no_bets)],
        ["Average edge", float(avg_edge) if pd.notna(avg_edge) else 0],
        ["Average control score", float(df["control_score"].mean()) if "control_score" in df else 0],
        ["Average chaos score", float(df["chaos_score"].mean()) if "chaos_score" in df else 0],
        [],
        ["Top 10 Value Picks", ""],
        ["Rank", "Match", "Pick", "Value Edge", "Recommendation"],
    ]
    top = recommended.sort_values("value_edge", ascending=False).head(10)
    for rank, (_, row) in enumerate(top.iterrows(), start=1):
        rows.append([
            rank,
            f"{row.get('home_team', '')} vs {row.get('away_team', '')}",
            row.get("value_pick", ""),
            row.get("value_edge", 0),
            row.get("bet_recommendation", ""),
        ])
    rows.extend([[], ["Most Common No-Bet Reasons", "Count"]])
    for reason, count in Counter(_split_reasons(df["no_bet_reasons"])).most_common(10):
        rows.append([reason, count])
    return rows


# ---------------------------------------------------------------------------
# Optional dashboard sheet builders
# ---------------------------------------------------------------------------

_METRIC_PERCENT_COLS = {"accuracy", "balanced_accuracy", "avg_confidence", "avg_correct_confidence"}
_METRIC_NUM_COLS = {"log_loss", "brier_score"}
_CALIBRATION_COLS = [
    "model_name", "calibrated", "log_loss", "brier_score",
    "avg_confidence", "avg_correct_confidence", "warnings",
]


def _add_model_comparison_sheet(wb: Workbook, csv_path: Path) -> None:
    df = pd.read_csv(csv_path)
    # Sort best first
    if "selected_as_best" in df.columns:
        df = df.sort_values("selected_as_best", ascending=False)
    ws = wb.create_sheet("Model Comparison")
    rows = [list(df.columns)] + df.where(pd.notna(df), None).values.tolist()
    _write_rows(ws, rows)
    _format_sheet(
        ws,
        percent_columns=_METRIC_PERCENT_COLS,
        odds_columns=set(),
        highlight_col="selected_as_best",
        highlight_fill=_BEST_FILL,
    )


def _add_best_model_sheet(wb: Workbook, meta_path: Path) -> None:
    try:
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
    except Exception:
        return
    ws = wb.create_sheet("Best Model")
    ws.append(["Field", "Value"])
    ws["A1"].font = _HEADER_FONT
    ws["A1"].fill = _HEADER_FILL
    ws["B1"].font = _HEADER_FONT
    ws["B1"].fill = _HEADER_FILL
    fields = [
        "model_name", "calibrated", "test_season", "selected_metric",
        "accuracy", "log_loss", "brier_score", "training_rows", "test_rows",
        "created_at",
    ]
    for field in fields:
        val = meta.get(field, "")
        ws.append([field, val])
    feature_cols = meta.get("feature_columns", [])
    if feature_cols:
        ws.append([])
        ws.append(["feature_columns", f"({len(feature_cols)} features)"])
        for fc in feature_cols:
            ws.append(["", fc])
    else:
        ws.append(["feature_columns", "WARNING: feature list not available"])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 48


def _add_backtest_sheet(wb: Workbook, bt_csv_path: Path) -> None:
    df = pd.read_csv(bt_csv_path)
    _add_dataframe_sheet(
        wb, "Backtest", df,
        percent_cols={"edge", "value_edge", "prob_home", "prob_draw", "prob_away"},
        odds_cols={"odds_home", "odds_draw", "odds_away", "odds_used"},
        highlight_col=None,
    )


def _add_backtest_summary_sheet(wb: Workbook, bt_csv_path: Path) -> None:
    df = pd.read_csv(bt_csv_path)
    ws = wb.create_sheet("Backtest Summary")
    ws.append(["Metric", "Value"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    total = len(df)
    bets_col = "bet_recommendation" if "bet_recommendation" in df.columns else None
    bet_mask = df[bets_col].fillna("").str.lower() != "no bet" if bets_col else pd.Series([True] * total)
    bets_df = df[bet_mask]
    no_bets_df = df[~bet_mask]
    n_bets = len(bets_df)
    n_no_bets = len(no_bets_df)

    profit_col = next((c for c in ("profit", "profit_units", "profit_stake") if c in df.columns), None)
    cum_col = next((c for c in ("cumulative_profit", "cumulative_profit_units") if c in df.columns), None)
    result_col = next((c for c in ("bet_result", "result") if c in df.columns), None)

    hit_rate: float | str = "n/a"
    if result_col and n_bets > 0:
        wins = bets_df[result_col].fillna("").str.lower().isin(["win", "w", "1"]).sum()
        hit_rate = float(wins / n_bets)

    total_profit: float | str = "n/a"
    roi: float | str = "n/a"
    best_profit: float | str = "n/a"
    worst_profit: float | str = "n/a"
    if profit_col and n_bets > 0:
        total_profit = float(bets_df[profit_col].sum())
        roi = float(total_profit / n_bets) if n_bets > 0 else 0.0
        best_profit = float(bets_df[profit_col].max())
        worst_profit = float(bets_df[profit_col].min())

    edge_col = next((c for c in ("value_edge", "edge") if c in df.columns), None)
    avg_edge: float | str = "n/a"
    if edge_col and n_bets > 0:
        avg_edge = float(bets_df[edge_col].mean())

    rows: list[list[object]] = [
        ["Total matches", total],
        ["Total bets placed", n_bets],
        ["No-bets", n_no_bets],
        ["Hit rate", hit_rate],
        ["Total profit (units)", total_profit],
        ["ROI per bet (units)", roi],
        ["Average value edge", avg_edge],
        ["Best single bet profit", best_profit],
        ["Worst single bet profit", worst_profit],
    ]
    for row in rows:
        ws.append(row)

    # No-bet reasons breakdown
    if "no_bet_reasons" in df.columns and n_no_bets > 0:
        ws.append([])
        ws.append(["Most Common No-Bet Reasons", "Count"])
        for cell in ws[ws.max_row]:
            cell.font = _BOLD
        for reason, count in Counter(_split_reasons(no_bets_df["no_bet_reasons"])).most_common(10):
            ws.append([reason, count])

    # Performance by recommendation
    if bets_col and profit_col and n_bets > 0:
        by_pick = bets_df.groupby(bets_col)[profit_col].agg(["count", "sum", "mean"]).reset_index()
        if not by_pick.empty:
            ws.append([])
            ws.append(["Performance by Recommendation", "Count", "Total Profit", "Avg Profit"])
            for cell in ws[ws.max_row]:
                cell.font = _BOLD
            for _, br in by_pick.iterrows():
                ws.append([br[bets_col], int(br["count"]), float(br["sum"]), float(br["mean"])])

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18


def _add_calibration_sheet(wb: Workbook, csv_path: Path) -> None:
    df = pd.read_csv(csv_path)
    ws = wb.create_sheet("Calibration")
    show_cols = [c for c in _CALIBRATION_COLS if c in df.columns]
    if not show_cols:
        ws.append(["No calibration data available."])
        return

    display_df = df[show_cols].copy()

    # Add interpretation column
    if "avg_confidence" in display_df.columns and "avg_correct_confidence" in display_df.columns:
        interps = []
        for _, row in display_df.iterrows():
            conf = row.get("avg_confidence")
            correct = row.get("avg_correct_confidence")
            if pd.isna(conf) or pd.isna(correct):
                interps.append("n/a")
            elif abs(conf - correct) < 0.05:
                interps.append("OK")
            elif conf > correct + 0.05:
                interps.append("Possibly overconfident")
            else:
                interps.append("Possibly underconfident")
        display_df = display_df.copy()
        display_df["calibration_interpretation"] = interps

    if "selected_as_best" in df.columns:
        display_df = display_df.copy()
        display_df["selected_as_best"] = df["selected_as_best"]
        display_df = display_df.sort_values("selected_as_best", ascending=False)

    rows = [list(display_df.columns)] + display_df.where(pd.notna(display_df), None).values.tolist()
    _write_rows(ws, rows)
    _format_sheet(
        ws,
        percent_columns=_METRIC_PERCENT_COLS,
        odds_columns=set(),
        highlight_col="selected_as_best",
        highlight_fill=_BEST_FILL,
    )


def _add_feature_metadata_sheet(wb: Workbook, meta_path: Path) -> None:
    try:
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
    except Exception:
        return
    ws = wb.create_sheet("Feature Metadata")
    ws.append(["Field", "Value"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    # Model metadata block
    ws.append(["--- Model Info ---", ""])
    for field in ["model_name", "calibrated", "test_season", "selected_metric",
                  "accuracy", "log_loss", "brier_score", "training_rows", "test_rows", "created_at"]:
        ws.append([field, meta.get(field, "")])

    # Feature columns block
    feature_cols = meta.get("feature_columns", [])
    ws.append([])
    if feature_cols:
        ws.append([f"Feature Columns ({len(feature_cols)} total)", ""])
        for fc in feature_cols:
            ws.append(["", fc])
    else:
        ws.append(["WARNING", "feature_columns list not available in metadata"])

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def create_predictions_excel_report(
    predictions_path: str | Path,
    output_path: str | Path,
    model_comparison_csv: str | Path | None = None,
    model_metadata_json: str | Path | None = None,
    backtest_csv: str | Path | None = None,
    backtest_report_md: str | Path | None = None,  # kept for signature compat, not used yet
) -> Path:
    """Create an Excel workbook with prediction results and optional dashboard sheets.

    The base sheets (Summary, Predictions, Value Bets, No Bets, High Chaos, v19 Flags)
    are always created. Optional dashboard sheets are added when the corresponding
    files are provided and exist.

    Args:
        predictions_path: CSV of predictions (required).
        output_path: Where to save the .xlsx file (required).
        model_comparison_csv: Optional path to model_comparison.csv.
        model_metadata_json: Optional path to best_model_metadata.json.
        backtest_csv: Optional path to backtest_bets.csv.
        backtest_report_md: Reserved for future use; currently unused.

    Returns:
        The output path.
    """
    predictions_path = Path(predictions_path)
    output_path = Path(output_path)
    df = pd.read_csv(predictions_path)
    data = _ensure_columns(df, PREDICTION_COLUMNS)

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    _write_rows(summary, _summary_rows(data))
    _format_sheet(summary, {"Value Edge", "Average edge"}, set())
    summary["A1"].font = Font(bold=True, color="FFFFFF")
    summary["B1"].font = Font(bold=True, color="FFFFFF")

    _add_dataframe_sheet(wb, "Predictions", data)
    value_bets = data[data["bet_recommendation"].fillna("").str.lower() != "no bet"].sort_values("value_edge", ascending=False)
    _add_dataframe_sheet(wb, "Value Bets", value_bets)
    no_bets = data[data["bet_recommendation"].fillna("").str.lower() == "no bet"]
    _add_dataframe_sheet(wb, "No Bets", no_bets)
    high_chaos = data.sort_values("chaos_score", ascending=False)
    _add_dataframe_sheet(wb, "High Chaos", high_chaos)
    flags = data[data["v19_flags"].fillna("").astype(str).str.len() > 0]
    _add_dataframe_sheet(wb, "v19 Flags", flags)

    # ---- Optional dashboard sheets ------------------------------------------
    cmp_csv = Path(model_comparison_csv) if model_comparison_csv else None
    meta_json = Path(model_metadata_json) if model_metadata_json else None
    bt_csv = Path(backtest_csv) if backtest_csv else None

    if cmp_csv and cmp_csv.exists():
        _add_model_comparison_sheet(wb, cmp_csv)
        _add_calibration_sheet(wb, cmp_csv)

    if meta_json and meta_json.exists():
        _add_best_model_sheet(wb, meta_json)
        _add_feature_metadata_sheet(wb, meta_json)  # sheet name uses no special chars

    if bt_csv and bt_csv.exists():
        _add_backtest_sheet(wb, bt_csv)
        _add_backtest_summary_sheet(wb, bt_csv)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
