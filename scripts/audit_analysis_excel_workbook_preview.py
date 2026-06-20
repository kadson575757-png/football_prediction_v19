# -*- coding: utf-8 -*-
"""Audit Phase 14.2 analysis Excel workbook previews."""
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]

ANALYSIS_EXCEL_WORKBOOK_PREVIEW_READY = "ANALYSIS_EXCEL_WORKBOOK_PREVIEW_READY"
BUILD_ANALYSIS_EXCEL_WORKBOOK_PREVIEW = "BUILD_ANALYSIS_EXCEL_WORKBOOK_PREVIEW"
FIX_ANALYSIS_EXCEL_WORKBOOK_PREVIEW = "FIX_ANALYSIS_EXCEL_WORKBOOK_PREVIEW"
INSTALL_OPENPYXL_OR_SKIP_EXCEL_PREVIEW = "INSTALL_OPENPYXL_OR_SKIP_EXCEL_PREVIEW"

XG_MODEL_INTEGRATION_NOT_ACTIVE_BY_DESIGN = "XG_MODEL_INTEGRATION_NOT_ACTIVE_BY_DESIGN"

OUTPUT_CSV = "analysis_excel_workbook_preview_summary.csv"
OUTPUT_MD = "analysis_excel_workbook_preview_summary.md"

EXPECTED_SHEETS = [
    "README",
    "Bundle_Index",
    "Match_Level",
    "Team_Aggregates",
    "Rolling_Form",
    "Matchup_Preview",
    "Reporting_Pack",
]

KEY_DATA_SHEETS = ["Bundle_Index", "Match_Level", "Team_Aggregates", "Rolling_Form", "Matchup_Preview", "Reporting_Pack"]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", default=None)
    parser.add_argument("--preview-dir", default=str(ROOT / "outputs" / "analysis_export_preview"))
    parser.add_argument("--output-dir", default=str(ROOT / "outputs" / "diagnostics"))
    parser.add_argument("--base-dir", default=str(ROOT))
    return parser


def _find_workbooks(workbook: str | Path | None, preview_dir: str | Path) -> list[Path]:
    if workbook:
        return [Path(workbook)]
    root = Path(preview_dir)
    return sorted(root.glob("*/analysis_export_workbook_preview.xlsx")) if root.exists() else []


def _under_export_preview(path_text: str, base: Path) -> bool:
    path = Path(path_text)
    if not path.is_absolute():
        path = base / path
    try:
        resolved = path.resolve()
    except OSError:
        return False
    allowed = (base / "outputs" / "analysis_export_preview").resolve()
    return resolved == allowed or allowed in resolved.parents


def _forbidden_path(path_text: str) -> bool:
    normalized = str(path_text).replace("\\", "/").lower()
    forbidden = [
        "/data/processed/",
        "/data/templates/",
        "/data/trusted_xg_sources/accepted/",
        "/data/trusted_xg_sources/raw/",
    ]
    return any(fragment in normalized for fragment in forbidden)


def _sheet_data_rows(ws: Any) -> int:
    return max(int(ws.max_row) - 1, 0)


def audit_workbook(path: Path, *, base_dir: str | Path = ROOT) -> dict[str, Any]:
    base = Path(base_dir).resolve()
    errors: list[str] = []
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - dependency presence varies by env
        return {
            "workbook_path": str(path),
            "sheets_found": 0,
            "missing_sheets": "",
            "zero_row_sheets": "",
            "model_integration_status": "",
            "workbook_valid": False,
            "blocking_reasons": f"MISSING_OPENPYXL:{exc}",
        }
    if not path.exists():
        return {
            "workbook_path": str(path),
            "sheets_found": 0,
            "missing_sheets": "ALL",
            "zero_row_sheets": "",
            "model_integration_status": "",
            "workbook_valid": False,
            "blocking_reasons": "WORKBOOK_NOT_FOUND",
        }
    if not _under_export_preview(str(path), base):
        errors.append("UNSAFE_WORKBOOK_PATH")
    if _forbidden_path(str(path)):
        errors.append("FORBIDDEN_PRODUCTION_WORKBOOK_PATH")
    wb = load_workbook(path, read_only=True, data_only=True)
    names = set(wb.sheetnames)
    missing = [sheet for sheet in EXPECTED_SHEETS if sheet not in names]
    if missing:
        errors.append("MISSING_EXPECTED_SHEETS")
    zero_rows: list[str] = []
    for sheet in KEY_DATA_SHEETS:
        if sheet in names and _sheet_data_rows(wb[sheet]) <= 0:
            zero_rows.append(sheet)
    if zero_rows:
        errors.append("ZERO_ROW_SHEET")
    model_status = ""
    if "README" in names:
        values = [str(cell.value) for row in wb["README"].iter_rows() for cell in row if cell.value is not None]
        joined = " | ".join(values)
        if XG_MODEL_INTEGRATION_NOT_ACTIVE_BY_DESIGN in joined:
            model_status = XG_MODEL_INTEGRATION_NOT_ACTIVE_BY_DESIGN
        else:
            errors.append("README_MISSING_MODEL_INTEGRATION_STATUS")
    return {
        "workbook_path": str(path),
        "sheets_found": int(len(wb.sheetnames)),
        "missing_sheets": " | ".join(missing),
        "zero_row_sheets": " | ".join(zero_rows),
        "model_integration_status": model_status,
        "workbook_valid": not errors,
        "blocking_reasons": " | ".join(errors),
    }


def recommendation(table: pd.DataFrame) -> str:
    if table.empty:
        return BUILD_ANALYSIS_EXCEL_WORKBOOK_PREVIEW
    if table["blocking_reasons"].astype(str).str.contains("MISSING_OPENPYXL").any():
        return INSTALL_OPENPYXL_OR_SKIP_EXCEL_PREVIEW
    if table["workbook_valid"].any():
        return ANALYSIS_EXCEL_WORKBOOK_PREVIEW_READY
    return FIX_ANALYSIS_EXCEL_WORKBOOK_PREVIEW


def build_markdown(table: pd.DataFrame, rec: str) -> str:
    lines = [
        "# Phase 14.2 Analysis Excel Workbook Preview Audit",
        "",
        "Phase 14.2 is an Excel/export/reporting preview only. xG remains inactive in model logic.",
        "",
        "## A. Executive Summary",
        f"- workbooks audited: {len(table)}",
        f"- valid workbooks: {int(table['workbook_valid'].sum()) if not table.empty else 0}",
        "",
        "## B. Workbook Diagnostics",
    ]
    if table.empty:
        lines += ["No analysis Excel workbook previews found.", ""]
    else:
        cols = ["sheets_found", "missing_sheets", "zero_row_sheets", "model_integration_status", "workbook_valid", "blocking_reasons"]
        lines += ["| " + " | ".join(cols) + " |", "| " + " | ".join("---" for _ in cols) + " |"]
        for _, row in table[cols].iterrows():
            lines.append("| " + " | ".join(str(row[col]).replace("|", ";") for col in cols) + " |")
        lines.append("")
    lines += [
        "## C. Safety Checks",
        "- Workbook path must stay under outputs/analysis_export_preview.",
        "- Workbook path must not point to production targets, manifests, accepted artifacts, or raw trusted sources.",
        "- No xG values inferred or invented.",
        "- No model, probability, market-tier, recommended-market, betting, staking, ROI, or SUPER_A_TIER logic changed.",
        "",
        "## D. Phase 14.2 Recommendation",
        rec,
        "",
    ]
    return "\n".join(lines)


def run(
    *,
    workbook: str | Path | None = None,
    preview_dir: str | Path = ROOT / "outputs" / "analysis_export_preview",
    output_dir: str | Path = ROOT / "outputs" / "diagnostics",
    base_dir: str | Path = ROOT,
) -> tuple[pd.DataFrame, str, str]:
    rows = [audit_workbook(path, base_dir=base_dir) for path in _find_workbooks(workbook, preview_dir)]
    table = pd.DataFrame(rows)
    rec = recommendation(table)
    markdown = build_markdown(table, rec)
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    table.to_csv(out / OUTPUT_CSV, index=False)
    (out / OUTPUT_MD).write_text(markdown, encoding="utf-8")
    return table, markdown, rec


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    table, _markdown, rec = run(workbook=args.workbook, preview_dir=args.preview_dir, output_dir=args.output_dir, base_dir=args.base_dir)
    print(f"Wrote {len(table)} rows to {Path(args.output_dir) / OUTPUT_CSV}")
    print(rec)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
